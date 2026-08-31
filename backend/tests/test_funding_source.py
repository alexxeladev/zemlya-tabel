"""task_funding_source — источник финансирования премий и KPI.

Премия/KPI с указанным юрлицом-источником относится на затраты ЭТОГО юрлица
целиком, а каскад распределения делит ОСТАТОК начисления:

    база каскада = Итого начислено − Σ целевых
    итог юрлица  = доля из каскада + его целевые суммы
    Σ по юрлицам = Итого начислено   (всегда, ровно)

Главная проверка — что целевые вычтены из базы, а не добавлены сверх неё
(двойной счёт: ведомость перестала бы сходиться с начислением).
"""
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.positions import EmployeePosition
from app.models.production_calendars import ProductionCalendar
from app.models.schedules import Schedule
from app.models.timesheet_entries import TimesheetEntry
from app.services.payroll_statement import (
    build_targeted_funding,
    cascade_base_amount,
    effective_percent,
    merge_targeted,
)
from tests.conftest import get_token

MAY_BASIC = {"year": 2026, "months": [{"month": 5, "days": "3,4,10,11,17,18,24,25,31"}]}
MAY_WORKDAYS = [d for d in range(1, 32) if d not in (3, 4, 10, 11, 17, 18, 24, 25, 31)]


# ── Фикстуры ──────────────────────────────────────────────────────────────────

@pytest.fixture
def dept(db_session: Session) -> Department:
    d = Department(name="Стройдепартамент", code="SD", is_active=True)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def companies(db_session: Session) -> list[Company]:
    """[0] ЗМО (основная компания сотрудника), [1] Секьюрити, [2] Комфорт."""
    cs = [
        Company(code="ZMO", name="ЗМО", is_active=True, sort_order=1),
        Company(code="SEC", name="Секьюрити", is_active=True, sort_order=2),
        Company(code="KMF", name="Комфорт", is_active=True, sort_order=3),
    ]
    db_session.add_all(cs)
    db_session.commit()
    for c in cs:
        db_session.refresh(c)
    return cs


@pytest.fixture
def schedule(db_session: Session) -> Schedule:
    s = Schedule(name="5/2", hours_per_shift=8, schedule_type="standard", is_active=True)
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture
def admin(db_session: Session) -> Employee:
    emp = Employee(full_name="Funding Admin", email="fundadmin@example.com",
                   hashed_password=hash_password("admin123"), role="admin",
                   is_active=True, must_change_password=False, is_system_admin=True)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def worker(db_session: Session, dept, companies, schedule) -> Employee:
    """Оклад 80 000, основная компания ЗМО."""
    emp = Employee(full_name="Иванов Иван", tab_number="F-1", is_active=True,
                   rate=Decimal("80000"), schedule_id=schedule.id,
                   default_company_id=companies[0].id, department_id=dept.id)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def calendar(db_session: Session) -> ProductionCalendar:
    cal = ProductionCalendar(year=2026, data=MAY_BASIC, source="manual")
    db_session.add(cal)
    db_session.commit()
    return cal


def _hdr(client: TestClient) -> dict:
    return {"Authorization": f"Bearer {get_token(client, 'fundadmin@example.com', 'admin123')}"}


def _full_norm_entries(db: Session, emp_id: int, company_id: int) -> None:
    """Полная норма мая по одной компании → оклад начисляется целиком."""
    for d in MAY_WORKDAYS:
        db.add(TimesheetEntry(employee_id=emp_id, work_date=date(2026, 5, d),
                              company_id=company_id, hours=8))
    db.commit()


def _set_shares(client: TestClient, hdr: dict, emp_id: int, shares: list[tuple[int, str]]):
    return client.put(
        f"/api/employees/{emp_id}/company-shares",
        json={"shares": [{"company_id": cid, "percent": p} for cid, p in shares]},
        headers=hdr,
    )


def _add_adjustment(client: TestClient, hdr: dict, emp_id: int, kind: str,
                    amount: str, funding_company_id: int | None = None,
                    position_id: int | None = None, reason: str = "за объект А"):
    payload = {
        "employee_id": emp_id, "year": 2026, "month": 5, "kind": kind,
        "amount": amount, "reason": reason,
    }
    if funding_company_id is not None:
        payload["funding_company_id"] = funding_company_id
    if position_id is not None:
        payload["position_id"] = position_id
    return client.post("/api/timesheet/adjustments", json=payload, headers=hdr)


def _row(client: TestClient, hdr: dict, emp_id: int, position_id: int | None = None) -> dict:
    r = client.get("/api/timesheet/2026/5/statement", headers=hdr)
    assert r.status_code == 200, r.text
    rows = [x for x in r.json()["rows"] if x["employee_id"] == emp_id]
    if position_id is not None:
        rows = [x for x in rows if x["position_id"] == position_id]
    assert rows, "строка сотрудника не найдена в ведомости"
    return rows[0]


def _amounts(row: dict) -> dict[int, Decimal]:
    return {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}


# ── Unit: арифметика целевых сумм ─────────────────────────────────────────────

class TestTargetedFundingUnit:
    def test_base_is_accrued_minus_targeted(self):
        t = build_targeted_funding(
            [(2, "premium", Decimal("20000"))], Decimal("100000"), {2: "Секьюрити"}
        )
        assert t.total == Decimal("20000")
        assert cascade_base_amount(Decimal("100000"), t) == Decimal("80000")

    def test_no_items_changes_nothing(self):
        t = build_targeted_funding(None, Decimal("100000"), {})
        assert t.total == Decimal("0")
        assert t.note is None
        assert cascade_base_amount(Decimal("100000"), t) == Decimal("100000")
        assert merge_targeted({1: Decimal("100000")}, t) == {1: Decimal("100000")}

    def test_base_never_negative(self):
        """Целевых больше начисленного — база каскада 0, а не минус."""
        t = build_targeted_funding(
            [(2, "premium", Decimal("150000"))], Decimal("100000"), {2: "Секьюрити"}
        )
        assert t.exceeds_accrued is True
        assert cascade_base_amount(Decimal("100000"), t) == Decimal("0")

    def test_same_company_targeted_twice_is_summed(self):
        t = build_targeted_funding(
            [(2, "premium", Decimal("5000")), (2, "kpi", Decimal("3000"))],
            Decimal("100000"), {2: "Секьюрити"},
        )
        assert t.amounts == {2: Decimal("8000")}
        assert t.total == Decimal("8000")

    def test_merge_adds_targeted_to_cascade_share(self):
        t = build_targeted_funding(
            [(2, "premium", Decimal("20000"))], Decimal("100000"), {2: "Секьюрити"}
        )
        merged = merge_targeted({1: Decimal("40000"), 2: Decimal("40000")}, t)
        assert merged == {1: Decimal("40000"), 2: Decimal("60000")}
        assert sum(merged.values()) == Decimal("100000")

    def test_merge_introduces_company_absent_from_cascade(self):
        t = build_targeted_funding(
            [(3, "kpi", Decimal("10000"))], Decimal("100000"), {3: "Комфорт"}
        )
        merged = merge_targeted({1: Decimal("90000")}, t)
        assert merged == {1: Decimal("90000"), 3: Decimal("10000")}

    def test_note_names_kind_amount_and_company(self):
        t = build_targeted_funding(
            [(2, "premium", Decimal("20000"))], Decimal("100000"), {2: "Секьюрити"}
        )
        assert t.note == "включает целевую премию 20000 ₽ (Секьюрити)"

    def test_note_lists_every_targeted_item(self):
        t = build_targeted_funding(
            [(2, "premium", Decimal("20000")), (3, "kpi", Decimal("5000"))],
            Decimal("100000"), {2: "Секьюрити", 3: "Комфорт"},
        )
        assert "целевую премию 20000 ₽ (Секьюрити)" in t.note
        assert "целевой KPI 5000 ₽ (Комфорт)" in t.note

    def test_effective_percent_is_share_of_accrued(self):
        assert effective_percent(Decimal("60000"), Decimal("100000")) == Decimal("60.00")
        assert effective_percent(Decimal("40000"), Decimal("100000")) == Decimal("40.00")
        assert effective_percent(Decimal("1"), Decimal("0")) == Decimal("0")


# ── Проверочный пример из ТЗ ──────────────────────────────────────────────────

class TestTaskExample:
    """Оклад 80 000 + премия 20 000 (Секьюрити), каскад 50/50 →
    ЗМО 40 000 (40%), Секьюрити 60 000 (60%), сумма 100 000."""

    @pytest.fixture
    def prepared(self, client, admin, worker, companies, schedule, calendar, db_session):
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        assert _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ]).status_code == 200
        assert _add_adjustment(
            client, hdr, worker.id, "premium", "20000", companies[1].id
        ).status_code == 201
        return hdr

    def test_amounts(self, client, prepared, worker, companies):
        row = _row(client, prepared, worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("100000")
        amounts = _amounts(row)
        assert amounts[companies[0].id] == Decimal("40000")
        assert amounts[companies[1].id] == Decimal("60000")
        assert sum(amounts.values()) == Decimal("100000")

    def test_sum_equals_accrued(self, client, prepared, worker):
        row = _row(client, prepared, worker.id)
        assert Decimal(row["distribution_total"]) == Decimal(row["accrued_total"])

    def test_effective_percents_are_40_60(self, client, prepared, worker, companies):
        row = _row(client, prepared, worker.id)
        eff = {d["company_id"]: Decimal(d["effective_percent"]) for d in row["distribution"]}
        assert eff[companies[0].id] == Decimal("40.00")
        assert eff[companies[1].id] == Decimal("60.00")

    def test_cascade_percent_stays_as_configured(self, client, prepared, worker, companies):
        """В поле percent остаётся ЗАДАННЫЙ каскадом %, иначе правка в ведомости
        учла бы целевую сумму второй раз."""
        row = _row(client, prepared, worker.id)
        pct = {d["company_id"]: Decimal(d["percent"]) for d in row["distribution"]}
        assert pct[companies[0].id] == Decimal("50")
        assert pct[companies[1].id] == Decimal("50")

    def test_note_present(self, client, prepared, worker):
        row = _row(client, prepared, worker.id)
        assert row["targeted_note"] == "включает целевую премию 20000 ₽ (Секьюрити)"
        assert Decimal(row["targeted_total"]) == Decimal("20000")

    def test_not_double_counted(self, client, prepared, worker, companies):
        """Каскад НЕ применён к полной сумме: иначе было бы 50 000 + 70 000."""
        amounts = _amounts(_row(client, prepared, worker.id))
        assert amounts[companies[0].id] != Decimal("50000")
        assert sum(amounts.values()) != Decimal("120000")


# ── Регрессия: без источника всё как раньше ───────────────────────────────────

class TestWithoutFundingSource:
    def test_plain_premium_distributed_by_cascade(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Премия без источника — часть общей базы, каскад 50/50 → 50 000 / 50 000."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        _add_adjustment(client, hdr, worker.id, "premium", "20000")

        row = _row(client, hdr, worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("100000")
        amounts = _amounts(row)
        assert amounts[companies[0].id] == Decimal("50000")
        assert amounts[companies[1].id] == Decimal("50000")
        assert row["targeted_note"] is None
        assert Decimal(row["targeted_total"]) == Decimal("0")

    def test_no_adjustments_at_all(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        amounts = _amounts(_row(client, hdr, worker.id))
        assert amounts[companies[0].id] == Decimal("40000")
        assert amounts[companies[1].id] == Decimal("40000")

    def test_auto_by_hours_unchanged(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Нижний уровень каскада (авто по часам) без целевых не изменился."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        row = _row(client, hdr, worker.id)
        assert row["distribution_source"] == "hours"
        assert _amounts(row)[companies[0].id] == Decimal("80000")


# ── Целевые начисления в разных ситуациях ─────────────────────────────────────

class TestTargetedInStatement:
    def test_kpi_with_funding_source(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Источник работает и у KPI, не только у премии."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        _add_adjustment(client, hdr, worker.id, "kpi", "20000", companies[1].id)

        row = _row(client, hdr, worker.id)
        amounts = _amounts(row)
        assert amounts[companies[0].id] == Decimal("40000")
        assert amounts[companies[1].id] == Decimal("60000")
        assert "целевой KPI" in row["targeted_note"]

    def test_several_targeted_for_one_employee(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Премия 20 000 (Секьюрити) + KPI 10 000 (Комфорт) при каскаде 50/50:
        база 110 000 − 30 000 = 80 000 → 40 000 / 40 000, плюс целевые."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)
        _add_adjustment(client, hdr, worker.id, "kpi", "10000", companies[2].id)

        row = _row(client, hdr, worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("110000")
        amounts = _amounts(row)
        assert amounts[companies[0].id] == Decimal("40000")
        assert amounts[companies[1].id] == Decimal("60000")
        assert amounts[companies[2].id] == Decimal("10000")
        assert sum(amounts.values()) == Decimal("110000")

    def test_funding_company_outside_cascade_appears(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Компания-источник не участвует в обычном распределении — появляется
        в разбивке отдельной строкой со своей суммой и нулевым % каскада."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [(companies[0].id, "100")])
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[2].id)

        row = _row(client, hdr, worker.id)
        entry = next(d for d in row["distribution"] if d["company_id"] == companies[2].id)
        assert Decimal(entry["amount"]) == Decimal("20000")
        assert Decimal(entry["percent"]) == Decimal("0")
        assert Decimal(entry["effective_percent"]) == Decimal("20.00")
        assert _amounts(row)[companies[0].id] == Decimal("80000")
        assert sum(_amounts(row).values()) == Decimal("100000")

    def test_cascade_base_zero(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Всё начисленное — целевая премия (часов нет): распределение состоит
        только из неё, каскад делит ноль."""
        hdr = _hdr(client)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)

        row = _row(client, hdr, worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("20000")
        amounts = _amounts(row)
        assert amounts[companies[1].id] == Decimal("20000")
        assert amounts.get(companies[0].id, Decimal("0")) == Decimal("0")
        assert sum(amounts.values()) == Decimal("20000")

    def test_targeted_with_auto_by_hours(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Нижний уровень каскада тоже делит УМЕНЬШЕННУЮ базу."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)

        row = _row(client, hdr, worker.id)
        assert row["distribution_source"] == "hours"
        amounts = _amounts(row)
        assert amounts[companies[0].id] == Decimal("80000")
        assert amounts[companies[1].id] == Decimal("20000")
        assert sum(amounts.values()) == Decimal(row["accrued_total"])

    def test_targeted_with_monthly_override(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Верхний уровень каскада (правка на месяц) — то же правило."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)
        client.put("/api/timesheet/distribution", json={
            "employee_id": worker.id, "year": 2026, "month": 5,
            "shares": [{"company_id": companies[2].id, "percent": "100"}],
        }, headers=hdr)

        row = _row(client, hdr, worker.id)
        assert row["distribution_source"] == "month"
        amounts = _amounts(row)
        assert amounts[companies[2].id] == Decimal("80000")
        assert amounts[companies[1].id] == Decimal("20000")
        assert sum(amounts.values()) == Decimal("100000")

    def test_totals_row_matches_accrued(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Σ распределения по ВСЕЙ ведомости = Σ «Итого начислено»."""
        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)

        data = client.get("/api/timesheet/2026/5/statement", headers=hdr).json()
        dist_sum = sum(Decimal(v) for v in data["distribution_totals"].values())
        assert dist_sum == Decimal(data["total_accrued"])


# ── Совместительство: целевая сумма влияет на СВОЮ позицию ────────────────────

class TestTargetedAndPositions:
    @pytest.fixture
    def combiner(self, db_session, worker, companies, schedule, dept):
        """Вторая позиция (электрик, оклад 30 000) в том же отделе."""
        worker.positions.append(EmployeePosition(
            title="Электрик", rate=Decimal("30000"), schedule_id=schedule.id,
            department_id=dept.id, company_id=companies[2].id,
        ))
        db_session.commit()
        db_session.refresh(worker)
        return worker

    def test_targeted_affects_only_its_position(
        self, client, admin, combiner, companies, schedule, calendar, db_session
    ):
        hdr = _hdr(client)
        primary = combiner.primary_position
        second = next(p for p in combiner.positions if not p.is_primary)
        _add_adjustment(client, hdr, combiner.id, "premium", "20000",
                        companies[1].id, position_id=second.id)

        primary_row = _row(client, hdr, combiner.id, primary.id)
        second_row = _row(client, hdr, combiner.id, second.id)

        assert Decimal(primary_row["targeted_total"]) == Decimal("0")
        assert Decimal(second_row["targeted_total"]) == Decimal("20000")
        assert _amounts(second_row)[companies[1].id] == Decimal("20000")
        for row in (primary_row, second_row):
            assert sum(_amounts(row).values()) == Decimal(row["accrued_total"])


# ── API источника финансирования ──────────────────────────────────────────────

class TestFundingSourceApi:
    def test_stored_and_returned_with_company_name(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        hdr = _hdr(client)
        created = _add_adjustment(client, hdr, worker.id, "premium", "5000",
                                  companies[1].id)
        assert created.status_code == 201
        body = created.json()
        assert body["funding_company_id"] == companies[1].id
        assert body["funding_company_name"] == "Секьюрити"

        listed = client.get("/api/timesheet/2026/5/adjustments", headers=hdr).json()
        item = next(a for a in listed if a["id"] == body["id"])
        assert item["funding_company_id"] == companies[1].id
        assert item["funding_company_name"] == "Секьюрити"

    def test_optional_by_default(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        hdr = _hdr(client)
        body = _add_adjustment(client, hdr, worker.id, "premium", "5000").json()
        assert body["funding_company_id"] is None
        assert body["funding_company_name"] is None

    def test_advance_rejects_funding_source(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Аванс — удержание, затрат юрлица за ним нет."""
        hdr = _hdr(client)
        r = _add_adjustment(client, hdr, worker.id, "advance", "5000", companies[1].id)
        assert r.status_code == 422

    def test_unknown_company_rejected(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        hdr = _hdr(client)
        r = _add_adjustment(client, hdr, worker.id, "premium", "5000", 99999)
        assert r.status_code == 422


# ── Excel-выгрузка ────────────────────────────────────────────────────────────

class TestExcelExport:
    def test_excel_amounts_match_web(
        self, client, admin, worker, companies, schedule, calendar, db_session
    ):
        """Суммы в выгрузке — те же, что в вебе, и пометка о целевой премии есть."""
        from io import BytesIO

        from openpyxl import load_workbook

        hdr = _hdr(client)
        _full_norm_entries(db_session, worker.id, companies[0].id)
        _set_shares(client, hdr, worker.id, [
            (companies[0].id, "50"), (companies[1].id, "50"),
        ])
        _add_adjustment(client, hdr, worker.id, "premium", "20000", companies[1].id)

        web = _amounts(_row(client, hdr, worker.id))
        r = client.get("/api/timesheet/2026/5/statement/export/excel", headers=hdr)
        assert r.status_code == 200
        ws = load_workbook(BytesIO(r.content)).active

        cells = [c for row in ws.iter_rows() for c in row if c.value is not None]
        texts = [str(c.value) for c in cells]
        numbers = [Decimal(str(c.value)) for c in cells
                   if isinstance(c.value, (int, float))]
        for amount in web.values():
            assert amount in numbers, f"суммы {amount} нет в выгрузке"
        assert any("целевую премию" in t for t in texts)
