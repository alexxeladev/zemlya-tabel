"""Tests for task 3.11b: company % distribution + payroll statement."""
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.production_calendars import ProductionCalendar
from app.models.schedules import Schedule
from app.models.timesheet_entries import TimesheetEntry
from app.services.company_order import company_display_name
from app.services.distribution import (
    distribute,
    distribute_largest_remainder,
    split_equally,
)
from app.services.payroll_statement import EMPTY_TARGETED, finalize_distribution
from tests.conftest import get_token

MAY_BASIC = {"year": 2026, "months": [{"month": 5, "days": "3,4,10,11,17,18,24,25,31"}]}
MAY_WORKDAYS = [d for d in range(1, 32) if d not in (3, 4, 10, 11, 17, 18, 24, 25, 31)]


# ── Unit: распределение по процентам ──────────────────────────────────────────

def _by_percent(total, shares, main=None):
    """Суммы каскада без целевых премий — то, что считает ведомость."""
    return finalize_distribution(total, shares, EMPTY_TARGETED)


class TestDistributeByPercent:
    def test_example_from_task(self):
        """120000 при 50/30/20 → 60000 / 36000 / 24000."""
        result = _by_percent(
            Decimal("120000"), {1: Decimal("50"), 2: Decimal("30"), 3: Decimal("20")}
        )
        assert result[1] == Decimal("60000")
        assert result[2] == Decimal("36000")
        assert result[3] == Decimal("24000")
        assert sum(result.values()) == Decimal("120000")

    def test_sum_matches_total_with_rounding(self):
        """Доли, не делящиеся нацело, всё равно сходятся с итогом."""
        result = _by_percent(
            Decimal("100"), {1: Decimal("33.33"), 2: Decimal("33.33"), 3: Decimal("33.34")}
        )
        assert sum(result.values()) == Decimal("100")

    def test_normalizes_non_100_sum(self):
        """Сумма процентов ≠ 100 — всё равно распределяет всю сумму (нормализация)."""
        shares = {1: Decimal("50"), 2: Decimal("50"), 3: Decimal("50")}
        result = _by_percent(Decimal("100"), shares)
        assert sum(result.values()) == Decimal("100")

    def test_empty_shares(self):
        assert _by_percent(Decimal("1000"), {}) == {}


# ── Unit: единый модуль распределения (task_distribution_v2 ч.1) ───────────────

class TestDistributionRounding:
    def test_350000_on_six_companies(self):
        """Пример из ТЗ: 350000 на 6 равных долей → 5×58333 + основная 58335."""
        shares = {c: Decimal("16.67") for c in range(1, 7)}
        result = distribute(Decimal("350000"), shares, main_key=3)
        assert sum(result.values()) == Decimal("350000")
        assert result[3] == Decimal("58335")
        assert sorted(result.values()) == [Decimal("58333")] * 5 + [Decimal("58335")]

    def test_remainder_goes_to_main_company(self):
        shares = {1: Decimal("1"), 2: Decimal("1"), 3: Decimal("1")}
        result = distribute(Decimal("100"), shares, main_key=2)
        assert sum(result.values()) == Decimal("100")
        assert result[2] == Decimal("34")
        assert result[1] == result[3] == Decimal("33")

    def test_remainder_to_largest_share_when_main_absent(self):
        """Основная компания не входит в распределение → остаток компании с наибольшей долей."""
        shares = {1: Decimal("20"), 2: Decimal("50"), 3: Decimal("30")}
        result = distribute(Decimal("1001"), shares, main_key=99)
        assert sum(result.values()) == Decimal("1001")
        assert result[2] == Decimal("501")  # наибольшая доля забрала остаток
        assert result[1] == Decimal("200")
        assert result[3] == Decimal("300")

    def test_sum_exact_for_many_random_percents(self):
        shares = {i: Decimal("100") / 7 for i in range(1, 8)}
        for total in (Decimal("1"), Decimal("999"), Decimal("123457"), Decimal("350000")):
            result = distribute(total, shares, main_key=1)
            assert sum(result.values()) == total

    def test_statement_uses_the_thousand_rounding(self):
        """Ведомость (без целевых премий) складывается ровно из общего
        `distribute_largest_remainder` — своей арифметики у неё нет."""
        shares = {1: Decimal("16.67"), 2: Decimal("16.67"), 3: Decimal("16.67"),
                  4: Decimal("16.67"), 5: Decimal("16.67"), 6: Decimal("16.65")}
        assert (finalize_distribution(Decimal("350000"), shares, EMPTY_TARGETED)
                == distribute_largest_remainder(Decimal("350000"), shares))


class TestSplitEqually:
    def test_six_companies_sum_100(self):
        """Пример из ТЗ: 6 компаний → по 16.67, основная 16.65, сумма ровно 100."""
        result = split_equally([1, 2, 3, 4, 5, 6], main_key=6)
        assert sum(result.values()) == Decimal("100")
        assert result[6] == Decimal("16.65")
        assert result[1] == Decimal("16.67")

    def test_three_companies(self):
        result = split_equally([10, 20, 30], main_key=10)
        assert sum(result.values()) == Decimal("100")
        assert result[10] == Decimal("33.34")
        assert result[20] == result[30] == Decimal("33.33")

    def test_single_company_gets_all(self):
        assert split_equally([7]) == {7: Decimal("100")}

    def test_empty_selection(self):
        assert split_equally([]) == {}

    def test_duplicates_ignored(self):
        result = split_equally([1, 1, 2])
        assert len(result) == 2
        assert sum(result.values()) == Decimal("100")


# ── Integration fixtures ──────────────────────────────────────────────────────

@pytest.fixture
def dept(db_session: Session) -> Department:
    d = Department(name="Stmt Dept", code="SD", is_active=True)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def companies(db_session: Session) -> list[Company]:
    cs = [
        Company(code="KMF", name="Комфорт", is_active=True),
        Company(code="ZMO", name="ЗМО", is_active=True),
        Company(code="GHS", name="ГХС", is_active=True),
    ]
    db_session.add_all(cs)
    db_session.commit()
    for c in cs:
        db_session.refresh(c)
    return cs


@pytest.fixture
def schedule(db_session: Session) -> Schedule:
    s = Schedule(name="5/2", hours_per_shift=8, schedule_type="standard", is_active=True)
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture
def admin(db_session: Session) -> Employee:
    emp = Employee(full_name="Stmt Admin", email="stmtadmin@example.com",
                   hashed_password=hash_password("admin123"), role="admin",
                   is_active=True, must_change_password=False, is_system_admin=True)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def worker(db_session: Session, dept, companies, schedule) -> Employee:
    emp = Employee(full_name="Кладовщик", tab_number="K-1", is_active=True,
                   rate=Decimal("80000"), schedule_id=schedule.id,
                   default_company_id=companies[0].id, department_id=dept.id)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def calendar(db_session: Session) -> ProductionCalendar:
    cal = ProductionCalendar(year=2026, data=MAY_BASIC, source="manual")
    db_session.add(cal)
    db_session.commit()
    return cal


def _full_norm_entries(db: Session, emp_id: int, company_id: int):
    for d in MAY_WORKDAYS:
        db.add(TimesheetEntry(employee_id=emp_id, work_date=date(2026, 5, d),
                              company_id=company_id, hours=8))
    db.commit()


def _h(client, token):
    return {"Authorization": f"Bearer {token}"}


# ── Default shares in employee card ───────────────────────────────────────────

class TestDefaultShares:
    def test_set_and_get_shares(self, client: TestClient, admin, worker, companies):
        token = get_token(client, "stmtadmin@example.com", "admin123")
        hdr = _h(client, token)
        payload = {"shares": [
            {"company_id": companies[0].id, "percent": "50"},
            {"company_id": companies[1].id, "percent": "30"},
            {"company_id": companies[2].id, "percent": "20"},
        ]}
        r = client.put(f"/api/employees/{worker.id}/company-shares", json=payload, headers=hdr)
        assert r.status_code == 200
        data = r.json()
        assert Decimal(data["percent_sum"]) == Decimal("100")
        assert len(data["shares"]) == 3

        g = client.get(f"/api/employees/{worker.id}/company-shares", headers=hdr)
        assert g.status_code == 200
        assert len(g.json()["shares"]) == 3

    def test_reject_sum_not_100(self, client: TestClient, admin, worker, companies):
        token = get_token(client, "stmtadmin@example.com", "admin123")
        payload = {"shares": [
            {"company_id": companies[0].id, "percent": "50"},
            {"company_id": companies[1].id, "percent": "30"},
        ]}
        r = client.put(f"/api/employees/{worker.id}/company-shares", json=payload,
                       headers=_h(client, token))
        assert r.status_code == 422


# ── Statement endpoint ────────────────────────────────────────────────────────

class TestStatement:
    def test_example_distribution(self, client: TestClient, admin, worker, companies,
                                   schedule, calendar, db_session):
        """Итого начислено 80000, 50/30/20 → 40000 / 24000 / 16000."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "50"},
            {"company_id": companies[1].id, "percent": "30"},
            {"company_id": companies[2].id, "percent": "20"},
        ]}, headers=_h(client, token))

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        assert r.status_code == 200
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("80000")
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert amounts[companies[0].id] == Decimal("40000")
        assert amounts[companies[1].id] == Decimal("24000")
        assert amounts[companies[2].id] == Decimal("16000")
        assert Decimal(row["distribution_total"]) == Decimal("80000")

    def test_monthly_override(self, client: TestClient, admin, worker, companies,
                              schedule, calendar, db_session):
        """Переопределение на месяц меняет распределение, не трогая карточку."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "50"},
            {"company_id": companies[1].id, "percent": "50"},
        ]}, headers=_h(client, token))

        # override: всё на одну компанию
        ov = client.put("/api/timesheet/distribution", json={
            "employee_id": worker.id, "year": 2026, "month": 5,
            "shares": [{"company_id": companies[2].id, "percent": "100"}],
        }, headers=_h(client, token))
        assert ov.status_code == 200

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["is_overridden"] is True
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert amounts[companies[2].id] == Decimal("80000")

        # карточка не изменилась
        g = client.get(f"/api/employees/{worker.id}/company-shares", headers=_h(client, token))
        assert len(g.json()["shares"]) == 2

        # удалить override → вернётся дефолт
        d = client.delete(f"/api/timesheet/distribution/{worker.id}/2026/5",
                          headers=_h(client, token))
        assert d.status_code == 204
        r2 = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row2 = next(x for x in r2.json()["rows"] if x["employee_id"] == worker.id)
        assert row2["is_overridden"] is False

class TestAutoDistributionByHours:
    """Задача task_distribution_fix: если ручной % не задан — распределять
    автоматически по фактическим часам сотрудника из табеля."""

    def _two_company_entries(self, db: Session, emp_id, c0, c1):
        """167 ч на c0 (20×8 + 1×7) + 4 ч на c1 (1×4) = 171 ч всего."""
        wd = MAY_WORKDAYS
        for d in wd[:20]:
            db.add(TimesheetEntry(employee_id=emp_id, work_date=date(2026, 5, d),
                                  company_id=c0, hours=8))
        db.add(TimesheetEntry(employee_id=emp_id, work_date=date(2026, 5, wd[20]),
                              company_id=c0, hours=7))
        db.add(TimesheetEntry(employee_id=emp_id, work_date=date(2026, 5, wd[21]),
                              company_id=c1, hours=4))
        db.commit()

    def test_auto_by_hours_no_manual_percent(self, client, admin, worker, companies,
                                             schedule, calendar, db_session):
        """167 ksec + 4 rest, без ручных % → 97.66% / 2.34%, сумма = «К выплате».

        База распределения — «К выплате» (task_it_arm_distribution ч.2), а сами
        доли округлены до тысячи (ч.3): проценты справочные, суммы по ним не
        пересчитываются в лоб.
        """
        self._two_company_entries(db_session, worker.id, companies[0].id, companies[1].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)

        assert row["is_auto_distributed"] is True
        assert row["is_overridden"] is False
        pcts = {d["company_id"]: Decimal(d["percent"]) for d in row["distribution"]}
        assert pcts[companies[0].id] == Decimal("97.66")
        assert pcts[companies[1].id] == Decimal("2.34")

        net = Decimal(row["net_payout"])
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert sum(amounts.values()) == net
        assert Decimal(row["distribution_total"]) == net
        assert all(a % 1000 == 0 for a in amounts.values())
        # доля больших часов получает большую сумму
        assert amounts[companies[0].id] > amounts[companies[1].id]

    def test_manual_percent_overrides_hours(self, client, admin, worker, companies,
                                            schedule, calendar, db_session):
        """С ручными % распределение по ним, часы игнорируются (не авто)."""
        self._two_company_entries(db_session, worker.id, companies[0].id, companies[1].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "50"},
            {"company_id": companies[1].id, "percent": "30"},
            {"company_id": companies[2].id, "percent": "20"},
        ]}, headers=_h(client, token))

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["is_auto_distributed"] is False
        net = Decimal(row["net_payout"])
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        # 50% от «К выплате», округлённые до тысячи (вниз, с раздачей остатка).
        half = net * Decimal("50") / Decimal("100")
        assert abs(amounts[companies[0].id] - half) < Decimal("1000")
        assert amounts[companies[0].id] % 1000 == 0
        assert sum(amounts.values()) == net
        assert companies[2].id in amounts  # компания без часов, но с ручным %

    def test_clearing_manual_returns_to_auto(self, client, admin, worker, companies,
                                             schedule, calendar, db_session):
        """Override → ручное; удаление override и пустая карточка → снова авто по часам."""
        self._two_company_entries(db_session, worker.id, companies[0].id, companies[1].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        client.put("/api/timesheet/distribution", json={
            "employee_id": worker.id, "year": 2026, "month": 5,
            "shares": [{"company_id": companies[2].id, "percent": "100"}],
        }, headers=_h(client, token))
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["is_overridden"] is True
        assert row["is_auto_distributed"] is False

        client.delete(f"/api/timesheet/distribution/{worker.id}/2026/5", headers=_h(client, token))
        r2 = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row2 = next(x for x in r2.json()["rows"] if x["employee_id"] == worker.id)
        assert row2["is_overridden"] is False
        assert row2["is_auto_distributed"] is True
        pcts = {d["company_id"]: Decimal(d["percent"]) for d in row2["distribution"]}
        assert pcts[companies[0].id] == Decimal("97.66")

    def test_no_hours_no_percent_falls_back_to_main(self, client, admin, worker,
                                                    companies, schedule, calendar, db_session):
        """Нет часов и нет % → не падает, всё на основную компанию."""
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["is_auto_distributed"] is True
        assert len(row["distribution"]) == 1
        assert row["distribution"][0]["company_id"] == companies[0].id  # default_company

    def test_employee_forbidden(self, client: TestClient, db_session):
        emp = Employee(full_name="E", email="stmtemp@example.com",
                       hashed_password=hash_password("emp12345"), role="employee",
                       is_active=True, must_change_password=False)
        db_session.add(emp)
        db_session.commit()
        token = get_token(client, "stmtemp@example.com", "emp12345")
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        assert r.status_code == 403

    def test_excel_export(self, client: TestClient, admin, worker, companies,
                          schedule, calendar, db_session):
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "100"},
        ]}, headers=_h(client, token))
        r = client.get("/api/timesheet/2026/5/statement/export/excel", headers=_h(client, token))
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(r.content) > 0


# ── Обоснования премий/KPI/удержаний в ведомости (task_ux_improvements ч.1b) ──

class TestAdjustmentReasonsInStatement:
    """Обоснование заводится обязательным у каждой премии/KPI/аванса — оно должно
    доезжать до ведомости и её Excel-выгрузки, а не оставаться только в табеле."""

    def _add(self, client, token, emp_id, kind, amount, reason):
        return client.post("/api/timesheet/adjustments", json={
            "employee_id": emp_id, "year": 2026, "month": 5,
            "kind": kind, "amount": amount, "reason": reason,
        }, headers=_h(client, token))

    def test_reasons_in_statement_rows(self, client: TestClient, admin, worker,
                                       companies, schedule, calendar, db_session):
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        assert self._add(client, token, worker.id, "premium", "5000",
                         "за переработку в мае").status_code in (200, 201)
        assert self._add(client, token, worker.id, "kpi", "3000",
                         "выполнение плана").status_code in (200, 201)
        assert self._add(client, token, worker.id, "advance", "2000",
                         "аванс 20 мая").status_code in (200, 201)

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        assert r.status_code == 200
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["premium_reasons"] == ["5000 ₽ — за переработку в мае"]
        assert row["kpi_reasons"] == ["3000 ₽ — выполнение плана"]
        assert row["advance_reasons"] == ["2000 ₽ — аванс 20 мая"]
        # Займа нет — примечания по нему тоже
        assert row["loan_note"] is None

    def test_several_adjustments_of_one_kind_listed_separately(
        self, client: TestClient, admin, worker, companies, schedule, calendar, db_session,
    ):
        """Две премии за месяц суммируются в деньгах, но обоснования не склеиваются:
        одна строка «Премия 12000» без разбивки не объясняет цифру."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._add(client, token, worker.id, "premium", "5000", "за объект А")
        self._add(client, token, worker.id, "premium", "7000", "за объект Б")

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert Decimal(row["premium_amount"]) == Decimal("12000")
        assert row["premium_reasons"] == ["5000 ₽ — за объект А", "7000 ₽ — за объект Б"]

    def test_no_adjustments_gives_empty_lists(self, client: TestClient, admin, worker,
                                              companies, schedule, calendar, db_session):
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["premium_reasons"] == []
        assert row["kpi_reasons"] == []
        assert row["advance_reasons"] == []

    def test_reasons_reach_excel(self, client: TestClient, admin, worker, companies,
                                 schedule, calendar, db_session):
        from io import BytesIO

        from openpyxl import load_workbook

        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._add(client, token, worker.id, "premium", "5000", "за переработку в мае")
        self._add(client, token, worker.id, "kpi", "3000", "выполнение плана")
        self._add(client, token, worker.id, "advance", "2000", "аванс 20 мая")

        r = client.get("/api/timesheet/2026/5/statement/export/excel",
                       headers=_h(client, token))
        assert r.status_code == 200
        ws = load_workbook(BytesIO(r.content)).active
        texts = [
            str(c.value)
            for row in ws.iter_rows()
            for c in row
            if isinstance(c.value, str)
        ]
        assert any("Обоснование премии" in t for t in texts)
        assert any("за переработку в мае" in t for t in texts)
        assert any("выполнение плана" in t for t in texts)
        assert any("аванс 20 мая" in t for t in texts)


# ── Дефолт отдела и каскад приоритетов (task_distribution_v2 ч.3) ─────────────

class TestDepartmentDefaultShares:
    """Каскад: месячный % > карточка сотрудника > дефолт отдела > авто по часам."""

    def _set_dept_shares(self, client, token, dept_id, shares):
        return client.put(f"/api/departments/{dept_id}/company-shares",
                          json={"shares": shares}, headers=_h(client, token))

    def test_set_and_get_department_shares(self, client: TestClient, admin, dept, companies):
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[0].id, "percent": "60"},
            {"company_id": companies[1].id, "percent": "40"},
        ])
        assert r.status_code == 200
        assert Decimal(r.json()["percent_sum"]) == Decimal("100")

        g = client.get(f"/api/departments/{dept.id}/company-shares", headers=_h(client, token))
        assert g.status_code == 200
        assert len(g.json()["shares"]) == 2

    def test_reject_sum_not_100(self, client: TestClient, admin, dept, companies):
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[0].id, "percent": "70"},
        ])
        assert r.status_code == 422

    def test_employee_without_own_shares_inherits_department(
        self, client: TestClient, admin, worker, dept, companies, schedule, calendar, db_session
    ):
        """Нет своего распределения → берётся дефолт отдела (не авто по часам)."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[1].id, "percent": "75"},
            {"company_id": companies[2].id, "percent": "25"},
        ])

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["distribution_source"] == "department"
        assert row["is_auto_distributed"] is False
        assert row["is_overridden"] is False
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert amounts[companies[1].id] == Decimal("60000")
        assert amounts[companies[2].id] == Decimal("20000")
        assert sum(amounts.values()) == Decimal(row["accrued_total"])

    def test_own_shares_win_over_department(
        self, client: TestClient, admin, worker, dept, companies, schedule, calendar, db_session
    ):
        """Индивидуальное распределение перекрывает дефолт отдела."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[1].id, "percent": "100"},
        ])
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "100"},
        ]}, headers=_h(client, token))

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["distribution_source"] == "employee"
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert amounts[companies[0].id] == Decimal("80000")

    def test_month_override_wins_over_all(
        self, client: TestClient, admin, worker, dept, companies, schedule, calendar, db_session
    ):
        """Правка на месяц — верх каскада (выше карточки и отдела)."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[1].id, "percent": "100"},
        ])
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "100"},
        ]}, headers=_h(client, token))
        client.put("/api/timesheet/distribution", json={
            "employee_id": worker.id, "year": 2026, "month": 5,
            "shares": [{"company_id": companies[2].id, "percent": "100"}],
        }, headers=_h(client, token))

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["distribution_source"] == "month"
        amounts = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert amounts[companies[2].id] == Decimal("80000")

    def test_falls_back_to_hours_when_nothing_set(
        self, client: TestClient, admin, worker, dept, companies, schedule, calendar, db_session
    ):
        """Ни одного уровня каскада → авто по часам."""
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["distribution_source"] == "hours"
        assert row["is_auto_distributed"] is True

    def test_clearing_department_default_returns_to_hours(
        self, client: TestClient, admin, worker, dept, companies, schedule, calendar, db_session
    ):
        _full_norm_entries(db_session, worker.id, companies[0].id)
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[1].id, "percent": "100"},
        ])
        self._set_dept_shares(client, token, dept.id, [])

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert row["distribution_source"] == "hours"

    def test_card_shows_inheritance(
        self, client: TestClient, admin, worker, dept, companies
    ):
        """В карточке видно, что распределение наследуется от отдела."""
        token = get_token(client, "stmtadmin@example.com", "admin123")
        self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[1].id, "percent": "100"},
        ])
        g = client.get(f"/api/employees/{worker.id}/company-shares", headers=_h(client, token))
        data = g.json()
        assert data["inherits_department"] is True
        assert data["department_name"] == dept.name
        assert len(data["department_shares"]) == 1

        # Задали своё → наследование выключается
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": companies[0].id, "percent": "100"},
        ]}, headers=_h(client, token))
        g2 = client.get(f"/api/employees/{worker.id}/company-shares", headers=_h(client, token))
        assert g2.json()["inherits_department"] is False

    def test_only_admin_can_set_department_shares(
        self, client: TestClient, admin, dept, companies, db_session
    ):
        acc = Employee(full_name="Acc", email="stmtacc@example.com",
                       hashed_password=hash_password("acc12345"), role="accountant",
                       is_active=True, must_change_password=False)
        db_session.add(acc)
        db_session.commit()
        token = get_token(client, "stmtacc@example.com", "acc12345")
        r = self._set_dept_shares(client, token, dept.id, [
            {"company_id": companies[0].id, "percent": "100"},
        ])
        assert r.status_code == 403
        # читать бухгалтер может
        g = client.get(f"/api/departments/{dept.id}/company-shares", headers=_h(client, token))
        assert g.status_code == 200


class TestScreenMatchesExcel:
    """AC 1 и 6: 350000 на 6 компаний — сумма ровно 350000 и на экране, и в Excel.

    Суммы следуют ЗАФИКСИРОВАННЫМ процентам (5 × 16.67% + основная 16.65%), а не
    идеальной 1/6: иначе ₽ в строке не сходились бы с показанным рядом %.
    С task_it_arm_distribution ч.3 доли ещё и округлены до ТЫСЯЧИ: точные 58345
    и 58275 floor-ятся до 58000, а две недостающие тысячи уходят наибольшим
    хвостам (345 против 275 — то есть 16.67-процентным, в порядке юрлиц).
    Сумма частей при этом по-прежнему ровно равна базе.
    """

    def test_six_equal_shares_sum_exactly(
        self, client: TestClient, admin, worker, companies, schedule, calendar, db_session
    ):
        from io import BytesIO

        from openpyxl import load_workbook

        # Оклад 350000, полная норма → Итого начислено ровно 350000.
        worker.rate = Decimal("350000")
        db_session.commit()
        _full_norm_entries(db_session, worker.id, companies[0].id)

        extra = [Company(code=f"C{i}", name=f"Компания {i}", is_active=True) for i in range(3)]
        db_session.add_all(extra)
        db_session.commit()
        six = companies + extra
        token = get_token(client, "stmtadmin@example.com", "admin123")

        # «Разнести поровну» между 6 компаниями: 5 × 16.67 + основная 16.65 = 100
        shares = split_equally([c.id for c in six], main_key=worker.default_company_id)
        client.put(f"/api/employees/{worker.id}/company-shares", json={"shares": [
            {"company_id": cid, "percent": str(pct)} for cid, pct in shares.items()
        ]}, headers=_h(client, token))

        r = client.get("/api/timesheet/2026/5/statement", headers=_h(client, token))
        row = next(x for x in r.json()["rows"] if x["employee_id"] == worker.id)
        assert Decimal(row["accrued_total"]) == Decimal("350000")
        # Удержаний нет → «К выплате» = 350000, она же база распределения.
        assert Decimal(row["net_payout"]) == Decimal("350000")
        screen = {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}
        assert sum(screen.values()) == Decimal("350000")
        assert sorted(screen.values()) == [Decimal("58000")] * 4 + [Decimal("59000")] * 2
        assert all(a % 1000 == 0 for a in screen.values())
        assert len(screen) == 6

        # Excel считает те же суммы (единый источник распределения)
        x = client.get("/api/timesheet/2026/5/statement/export/excel", headers=_h(client, token))
        ws = load_workbook(BytesIO(x.content)).active
        # Раскладка листа — по образцу финдира (task_vedomost_format ч.3):
        # шапка 1–5, заголовки колонок в 7-й, итоги в 8-й, сотрудники с 9-й.
        # Колонки юрлиц подписаны НАЗВАНИЕМ, а не кодом (ч.2).
        header_row, first_data_row = 7, 9
        col_by_company = {
            c.id: idx
            for idx, cell in enumerate(ws[header_row], start=1)
            for c in six
            if cell.value == company_display_name(c)
        }
        excel = {
            cid: Decimal(str(ws.cell(row=first_data_row, column=col).value))
            for cid, col in col_by_company.items()
        }
        assert excel == screen
        assert sum(excel.values()) == Decimal("350000")
