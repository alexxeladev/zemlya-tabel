"""Колонки переработки в ведомости (task_overtime_columns).

Требование: «Кол-во переработки» и «Сумма ПЕРЕРАБОТКи» показывают ТЕ ЖЕ часы, из
которых считаются деньги, — сверхурочные ПЛЮС работу в выходные/праздники по
графику. «Начислено, оклад» остаётся только оплатой обычных часов.

Это перекладка сумм между колонками отображения: расчёт не меняется, «Итого
начислено» и «К выплате» обязаны остаться прежними у КАЖДОГО сотрудника —
отдельный класс проверяет это против самого расчёта (`/payroll`).
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.departments import Department
from app.models.employee_adjustments import EmployeeAdjustment
from app.models.employees import Employee
from app.models.positions import EmployeePosition
from app.models.production_calendars import ProductionCalendar
from app.models.schedules import Schedule
from app.models.timesheet_entries import TimesheetEntry
from tests.conftest import get_token

# Май 2026: 21 рабочий день (норма 168 ч при смене 8 ч), 10 выходных Сб/Вс.
# Июнь 2026: те же 21 рабочий день, но один нерабочий день — БУДНИЙ (12 июня),
# то есть праздник по эвристике `is_public_holiday`.
CAL_2026 = {
    "year": 2026,
    "months": [
        {"month": 5, "days": "2,3,9,10,16,17,23,24,30,31"},
        {"month": 6, "days": "6,7,12,13,14,20,21,27,28"},
    ],
}
MAY_WEEKEND = [2, 3, 9, 10, 16, 17, 23, 24, 30, 31]
MAY_WORKDAYS = [d for d in range(1, 32) if d not in MAY_WEEKEND]

_ZERO = Decimal("0")


# ── Фикстуры ──────────────────────────────────────────────────────────────────

@pytest.fixture
def companies(db_session: Session) -> list[Company]:
    cs = [
        Company(code="OTA", name="Овертайм Один", is_active=True, sort_order=1),
        Company(code="OTB", name="Овертайм Два", is_active=True, sort_order=2),
    ]
    db_session.add_all(cs)
    db_session.commit()
    for c in cs:
        db_session.refresh(c)
    return cs


@pytest.fixture
def dept(db_session: Session, companies) -> Department:
    d = Department(name="Отдел переработок", code="OTD", is_active=True,
                   head_company_id=companies[0].id)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def dept2(db_session: Session, companies) -> Department:
    d = Department(name="Отдел подработок", code="OTD2", is_active=True,
                   head_company_id=companies[1].id)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def schedule(db_session: Session) -> Schedule:
    s = Schedule(name="5/2", hours_per_shift=8, schedule_type="weekday", is_active=True)
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture
def calendar(db_session: Session) -> ProductionCalendar:
    cal = ProductionCalendar(year=2026, data=CAL_2026, source="manual")
    db_session.add(cal)
    db_session.commit()
    return cal


@pytest.fixture
def admin(db_session: Session) -> Employee:
    emp = Employee(full_name="Админ переработок", email="overtime@example.com",
                   hashed_password=hash_password("admin123"), role="admin",
                   is_active=True, must_change_password=False, is_system_admin=True)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


# ── Хелперы ───────────────────────────────────────────────────────────────────

def _worker(db: Session, name, tab, dept, company, schedule, **fields) -> Employee:
    emp = Employee(full_name=name, tab_number=tab, is_active=True,
                   schedule_id=schedule.id, default_company_id=company.id,
                   department_id=dept.id, **fields)
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


def _hours(db: Session, emp: Employee, company: Company, month: int,
           by_day: dict[int, int], position: EmployeePosition | None = None):
    pid = (position or emp.primary_position).id
    for day, h in by_day.items():
        db.add(TimesheetEntry(employee_id=emp.id, position_id=pid,
                              work_date=date(2026, month, day),
                              company_id=company.id, hours=h))
    db.commit()


def _adjustment(db: Session, emp: Employee, kind: str, amount: str, actor: Employee):
    db.add(EmployeeAdjustment(
        employee_id=emp.id, position_id=emp.primary_position.id,
        year=2026, month=5, kind=kind, amount=Decimal(amount),
        reason="проверочный пример", created_by_id=actor.id))
    db.commit()


def _h(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _token(client: TestClient) -> str:
    return get_token(client, "overtime@example.com", "admin123")


def _statement(client: TestClient, token: str, month: int = 5) -> dict:
    r = client.get(f"/api/timesheet/2026/{month}/statement", headers=_h(token))
    assert r.status_code == 200, r.text
    return r.json()


def _payroll(client: TestClient, token: str, month: int = 5) -> dict:
    r = client.get(f"/api/timesheet/2026/{month}/payroll", headers=_h(token))
    assert r.status_code == 200, r.text
    return r.json()


def _row(statement: dict, employee_id: int, position_id: int | None = None) -> dict:
    return next(r for r in statement["rows"]
                if r["employee_id"] == employee_id
                and (position_id is None or r["position_id"] == position_id))


def _deev_days() -> dict[int, int]:
    """Раскладка часов Деева: 192 ч по плановым дням + 45 ч в выходные.

    Плановых дней 21: один день 4 ч (недобор 4 ч до нормы), семь по 12 ч
    (28 ч сверхурочных) и тринадцать по 8 ч. Зачётных часов 164, сверхурочных 28.
    Выходных дней выхода 7 (45 ч) — итого 237 ч за 28 дней.
    """
    days = {1: 4}
    for d in MAY_WORKDAYS[1:8]:
        days[d] = 12
    for d in MAY_WORKDAYS[8:]:
        days[d] = 8
    for d in (2, 3, 9, 10, 16, 17):
        days[d] = 6
    days[23] = 9
    return days


@pytest.fixture
def deev(db_session: Session, admin, calendar, dept, companies, schedule) -> Employee:
    """Проверочный пример задачи: оклад 75 000, 5/2, коэффициент 1.5."""
    emp = _worker(db_session, "Деев Д.Д.", "OT-1", dept, companies[0], schedule,
                  rate=Decimal("75000"))
    _hours(db_session, emp, companies[0], 5, _deev_days())
    _adjustment(db_session, emp, "premium", "5000", admin)
    return emp


# ── Проверочный пример ────────────────────────────────────────────────────────

class TestDeevExample:
    def test_hours_and_amounts(self, client, db_session, deev):
        """ОБЯЗАТЕЛЬНЫЙ пример задачи целиком.

        237 ч = 164 обычных (73 214 ₽) + 28 сверхурочных (18 750 ₽)
        + 45 в выходные (30 134 ₽); плюс премия 5 000 → начислено 127 098 ₽.
        """
        row = _row(_statement(client, _token(client)), deev.id)

        assert Decimal(row["norm_hours"]) == Decimal("168")
        assert row["norm_days"] == 21
        assert Decimal(row["fact_hours"]) == Decimal("237")
        assert row["fact_days"] == 28

        assert Decimal(row["base_salary"]) == Decimal("73214")
        assert Decimal(row["overtime_hours"]) == Decimal("73")
        assert Decimal(row["overtime_amount"]) == Decimal("48884")
        assert Decimal(row["accrued_total"]) == Decimal("127098")

    def test_overtime_hours_are_the_categories_paid_at_a_coefficient(
        self, client, db_session, deev
    ):
        """Часы колонки — те же категории, что и сумма рядом: сверхурочные плюс
        выходные/праздничные по графику. Именно из них считаются деньги, поэтому
        число и сумма в соседних колонках обязаны сходиться."""
        token = _token(client)
        pay = next(p for p in _payroll(client, token)["employees"]
                   if p["employee_id"] == deev.id)
        row = _row(_statement(client, token), deev.id)

        assert Decimal(row["overtime_hours"]) == (
            Decimal(pay["overtime_hours"]) + Decimal(pay["off_schedule_hours"])
            + Decimal(pay["holiday_hours"])
        )
        assert Decimal(row["overtime_amount"]) == (
            Decimal(pay["overtime_amount"]) + Decimal(pay["off_schedule_amount"])
            + Decimal(pay["holiday_amount"])
        )

    def test_salary_column_has_no_weekend_pay(self, client, db_session, deev):
        """«Начислено, оклад» — только обычные часы, без полуторных выходных."""
        token = _token(client)
        pay = next(p for p in _payroll(client, token)["employees"]
                   if p["employee_id"] == deev.id)
        row = _row(_statement(client, token), deev.id)
        assert Decimal(row["base_salary"]) == Decimal(pay["base_amount"])
        assert Decimal(pay["off_schedule_amount"]) > _ZERO
        assert Decimal(row["base_salary"]) < (
            Decimal(pay["base_amount"]) + Decimal(pay["off_schedule_amount"])
        )

    def test_delta_is_not_shown(self, client, db_session, deev):
        """Дельта (237 − 168 = 69) в ведомости не показывается вовсе: она
        схлопывает переработку с недоработкой и в расчёте не участвует.

        Тест-сторож: поля у строки ведомости никогда не было, и появиться оно
        не должно.
        """
        row = _row(_statement(client, _token(client)), deev.id)
        assert "delta_hours" not in row


# ── Ничего не потеряно и не задвоено ──────────────────────────────────────────

@pytest.fixture
def crew(db_session: Session, admin, calendar, dept, dept2, companies, schedule):
    """Выборка на все типы оплаты плюс совместитель — на ней сверяются итоги.

    Итоги нельзя сравнить «до и после» внутри одного прогона, поэтому эталоном
    служит САМ РАСЧЁТ (`/payroll`): задача его не трогает, и ведомость обязана
    сходиться с ним до копейки.
    """
    people: dict[str, Employee] = {}

    # Окладник с недобором нормы, переработкой и выходами в выходные.
    people["salary"] = _worker(db_session, "Окладов О.О.", "OT-2", dept,
                               companies[0], schedule, rate=Decimal("60000"))
    _hours(db_session, people["salary"], companies[0], 5, _deev_days())

    # Окладник ровно по норме — без переработки и выходных вовсе.
    people["exact"] = _worker(db_session, "Ровнов Р.Р.", "OT-3", dept,
                              companies[0], schedule, rate=Decimal("50000"))
    _hours(db_session, people["exact"], companies[0], 5,
           {d: 8 for d in MAY_WORKDAYS})

    # Посменный: смены в выходные оплачиваются целой сменой по коэффициенту.
    people["shift"] = _worker(db_session, "Сменов С.С.", "OT-4", dept,
                              companies[0], schedule, pay_type="per_shift",
                              shift_rate=Decimal("4000"))
    _hours(db_session, people["shift"], companies[0], 5,
           {**{d: 8 for d in MAY_WORKDAYS}, 2: 8, 3: 10})

    # Почасовик: категорий «вне графика» и «праздничные» у него нет.
    people["hourly"] = _worker(db_session, "Часов Ч.Ч.", "OT-5", dept,
                               companies[0], schedule, pay_type="hourly",
                               hour_rate=Decimal("450"))
    _hours(db_session, people["hourly"], companies[0], 5,
           {**{d: 10 for d in MAY_WORKDAYS}, 9: 6})

    # Совместитель: два рабочих места в разных отделах и юрлицах.
    comb = _worker(db_session, "Совместов В.В.", "OT-6", dept, companies[0],
                   schedule, rate=Decimal("70000"))
    extra = EmployeePosition(
        employee_id=comb.id, title="Электрик", is_primary=False, is_active=True,
        pay_type="salary", rate=Decimal("30000"), schedule_id=schedule.id,
        department_id=dept2.id, company_id=companies[1].id)
    db_session.add(extra)
    db_session.commit()
    db_session.refresh(extra)
    _hours(db_session, comb, companies[0], 5, _deev_days())
    _hours(db_session, comb, companies[1], 5, {2: 8, 3: 8, 4: 12}, position=extra)
    people["combiner"] = comb

    for kind, amount in (("premium", "3000"), ("kpi", "2000"), ("advance", "1500")):
        _adjustment(db_session, people["salary"], kind, amount, admin)
    return people


class TestNothingLostOrDuplicated:
    def test_accrued_and_net_payout_match_the_calculation(
        self, client, db_session, deev, crew
    ):
        """«Итого начислено» и «К выплате» не изменились НИ У КОГО.

        Перекладка сумм между колонками не имеет права трогать итог: сверяем
        каждую строку ведомости с расчётом, который задача не меняла.
        """
        token = _token(client)
        pays = {(p["employee_id"], p["position_id"]): p
                for p in _payroll(client, token)["employees"]}
        rows = _statement(client, token)["rows"]
        assert len(rows) >= 6

        for row in rows:
            p = pays[(row["employee_id"], row["position_id"])]
            expected_accrued = (Decimal(p["total_amount"])
                                + Decimal(p["premium_amount"])
                                + Decimal(p["kpi_amount"]))
            assert Decimal(row["accrued_total"]) == expected_accrued, row["employee_name"]
            assert Decimal(row["net_payout"]) == Decimal(p["net_payout"]), \
                row["employee_name"]

    def test_salary_and_overtime_columns_split_the_same_money(
        self, client, db_session, deev, crew
    ):
        """Сумма двух колонок = все рабочие категории расчёта: ни рубля мимо."""
        token = _token(client)
        pays = {(p["employee_id"], p["position_id"]): p
                for p in _payroll(client, token)["employees"]}

        for row in _statement(client, token)["rows"]:
            p = pays[(row["employee_id"], row["position_id"])]
            assert Decimal(row["base_salary"]) + Decimal(row["overtime_amount"]) == (
                Decimal(p["base_amount"]) + Decimal(p["overtime_amount"])
                + Decimal(p["off_schedule_amount"]) + Decimal(p["holiday_amount"])
            ), row["employee_name"]

    def test_statement_totals_still_add_up(self, client, db_session, deev, crew):
        """Подвал ведомости: Σ оклада и Σ переработки по-прежнему дают то же
        начислено вместе с остальными слагаемыми."""
        st = _statement(client, _token(client))
        assert Decimal(st["total_accrued"]) == sum(
            (Decimal(r["accrued_total"]) for r in st["rows"]), _ZERO
        )
        assert (Decimal(st["total_base_salary"])
                + Decimal(st["total_overtime_amount"])
                + Decimal(st["total_vacation_amount"])
                + Decimal(st["total_sick_amount"])
                + Decimal(st["total_night_amount"])
                + Decimal(st["total_premium"])
                + Decimal(st["total_kpi"])) == Decimal(st["total_accrued"])

    def test_distribution_is_untouched(self, client, db_session, deev, crew):
        """Распределение по юрлицам считается от «Итого начислено» и от
        перекладки колонок не зависит."""
        token = _token(client)
        for row in _statement(client, token)["rows"]:
            total = sum((Decimal(d["amount"]) for d in row["distribution"]), _ZERO)
            assert total + Decimal(row["unallocated_remainder"]) == \
                Decimal(row["accrued_total"]), row["employee_name"]


class TestPayTypes:
    def test_per_shift_weekend_shifts_go_to_the_overtime_column(
        self, client, db_session, crew
    ):
        """Посменный: смена в выходной оплачена по коэффициенту — её место в
        колонке переработки, а не в «Начислено, оклад»."""
        token = _token(client)
        emp = crew["shift"]
        p = next(x for x in _payroll(client, token)["employees"]
                 if x["employee_id"] == emp.id)
        row = _row(_statement(client, token), emp.id)

        assert Decimal(p["off_schedule_hours"]) > _ZERO
        assert Decimal(row["base_salary"]) == Decimal(p["base_amount"])
        assert Decimal(row["overtime_amount"]) == (
            Decimal(p["overtime_amount"]) + Decimal(p["off_schedule_amount"])
        )
        assert Decimal(row["overtime_hours"]) == (
            Decimal(p["overtime_hours"]) + Decimal(p["off_schedule_hours"])
        )

    def test_hourly_has_no_extra_categories(self, client, db_session, crew):
        """Почасовик: категорий «вне графика» и «праздничные» у него нет, поэтому
        колонка переработки равна его переработке, а оклад — фактическим часам."""
        token = _token(client)
        emp = crew["hourly"]
        p = next(x for x in _payroll(client, token)["employees"]
                 if x["employee_id"] == emp.id)
        row = _row(_statement(client, token), emp.id)

        assert Decimal(p["off_schedule_hours"]) == _ZERO
        assert Decimal(row["overtime_hours"]) == Decimal(p["overtime_hours"])
        assert Decimal(row["overtime_amount"]) == Decimal(p["overtime_amount"])
        assert Decimal(row["base_salary"]) == Decimal(p["base_amount"])

    def test_combiner_positions_are_split_independently(
        self, client, db_session, crew
    ):
        """У совместителя раскладка своя на каждом рабочем месте."""
        token = _token(client)
        emp = crew["combiner"]
        pays = {p["position_id"]: p for p in _payroll(client, token)["employees"]
                if p["employee_id"] == emp.id}
        rows = {r["position_id"]: r for r in _statement(client, token)["rows"]
                if r["employee_id"] == emp.id}
        assert len(rows) == 2

        for pid, row in rows.items():
            p = pays[pid]
            assert Decimal(row["base_salary"]) == Decimal(p["base_amount"])
            assert Decimal(row["overtime_amount"]) == (
                Decimal(p["overtime_amount"]) + Decimal(p["off_schedule_amount"])
                + Decimal(p["holiday_amount"])
            )


class TestHolidayWork:
    def test_holiday_pay_goes_to_the_overtime_column(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Праздничные часы — тоже повышенная категория: в колонку переработки,
        а не в оклад (12 июня 2026 — нерабочий будний день календаря)."""
        emp = _worker(db_session, "Праздников П.П.", "OT-7", dept, companies[0],
                      schedule, rate=Decimal("60000"))
        june_workdays = [d for d in range(1, 31)
                         if d not in (6, 7, 12, 13, 14, 20, 21, 27, 28)]
        _hours(db_session, emp, companies[0], 6,
               {**{d: 8 for d in june_workdays}, 12: 8})

        token = _token(client)
        p = next(x for x in _payroll(client, token, 6)["employees"]
                 if x["employee_id"] == emp.id)
        row = _row(_statement(client, token, 6), emp.id)

        assert Decimal(p["holiday_hours"]) == Decimal("8")
        assert Decimal(row["overtime_hours"]) == Decimal("8")
        assert Decimal(row["overtime_amount"]) == Decimal(p["holiday_amount"])
        assert Decimal(row["base_salary"]) == Decimal(p["base_amount"])


class TestExcelMatchesWeb:
    def test_deev_row_and_totals_match_the_web_statement(
        self, client, db_session, deev
    ):
        """Excel-выгрузка совпадает с веб-ведомостью: те же колонки, те же числа."""
        token = _token(client)
        row = _row(_statement(client, token), deev.id)

        resp = client.get("/api/timesheet/2026/5/statement/export/excel",
                          headers=_h(token))
        assert resp.status_code == 200
        ws = load_workbook(BytesIO(resp.content)).active

        header_row = 7
        headers = {ws.cell(header_row, c).value: c
                   for c in range(1, ws.max_column + 1)}
        data_row = next(r for r in range(header_row + 1, ws.max_row + 1)
                        if ws.cell(r, headers["ФИО"]).value == deev.full_name)

        def cell(title):
            return Decimal(str(ws.cell(data_row, headers[title]).value))

        assert cell("Кол-во переработки, час") == Decimal("73")
        assert cell("Сумма ПЕРЕРАБОТКи") == Decimal("48884")
        assert cell("Начислено, оклад") == Decimal("73214")
        assert cell("Итого начислено") == Decimal("127098")

        assert cell("Кол-во переработки, час") == Decimal(row["overtime_hours"])
        assert cell("Сумма ПЕРЕРАБОТКи") == Decimal(row["overtime_amount"])
        assert cell("Начислено, оклад") == Decimal(row["base_salary"])

        totals_row = header_row + 1
        assert Decimal(str(ws.cell(totals_row, headers["Сумма ПЕРЕРАБОТКи"]).value)) \
            == Decimal("48884")
        assert Decimal(str(ws.cell(totals_row, headers["Начислено, оклад"]).value)) \
            == Decimal("73214")
