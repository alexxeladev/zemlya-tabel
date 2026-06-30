"""Tests for T-13 Excel export."""
from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.production_calendars import ProductionCalendar
from app.models.timesheet_entries import TimesheetEntry
from tests.conftest import get_token

YEAR = 2026
MONTH = 6  # июнь — 30 дней


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def dept(db_session: Session) -> Department:
    d = Department(name="Отдел Тест", code="ОТ", is_active=True)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def company_alpha(db_session: Session) -> Company:
    c = Company(code="АЛФ", name="ООО Альфа", is_active=True)
    db_session.add(c)
    db_session.commit()
    db_session.refresh(c)
    return c


@pytest.fixture
def company_beta(db_session: Session) -> Company:
    c = Company(code="БЕТ", name="ООО Бета", is_active=True)
    db_session.add(c)
    db_session.commit()
    db_session.refresh(c)
    return c


@pytest.fixture
def admin_emp(db_session: Session, dept: Department) -> Employee:
    emp = Employee(
        full_name="Главный Администратор",
        email="admin@example.com",
        hashed_password=hash_password("admin123"),
        role="admin",
        is_active=True,
        is_system_admin=True,
        department_id=dept.id,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def employee_one(db_session: Session, dept: Department) -> Employee:
    emp = Employee(
        full_name="Иванов Иван Иванович",
        position="Инженер",
        tab_number="0001",
        email="emp1@example.com",
        hashed_password=hash_password("pass"),
        role="employee",
        is_active=True,
        department_id=dept.id,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def employee_two(db_session: Session, dept: Department) -> Employee:
    emp = Employee(
        full_name="Петров Пётр Петрович",
        position="Бухгалтер",
        tab_number="0002",
        email="emp2@example.com",
        hashed_password=hash_password("pass"),
        role="employee",
        is_active=True,
        department_id=dept.id,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def manager_emp(db_session: Session, dept: Department) -> Employee:
    emp = Employee(
        full_name="Менеджер Менеджерович",
        email="manager@example.com",
        hashed_password=hash_password("pass"),
        role="manager",
        is_active=True,
        department_id=dept.id,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def worker_emp(db_session: Session) -> Employee:
    """Сотрудник-работник без роли менеджера (employee)."""
    emp = Employee(
        full_name="Работник Работников",
        email="worker@example.com",
        hashed_password=hash_password("pass"),
        role="employee",
        is_active=True,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def entries_for_emp1(
    db_session: Session,
    employee_one: Employee,
    company_alpha: Company,
    company_beta: Company,
) -> list[TimesheetEntry]:
    """Иванов: 8ч в Альфа на рабочие дни 1-5, 4ч в Бета на день 2."""
    created = []
    for day in [1, 2, 3, 4, 5]:
        e = TimesheetEntry(
            employee_id=employee_one.id,
            work_date=date(YEAR, MONTH, day),
            company_id=company_alpha.id,
            hours=8,
        )
        db_session.add(e)
        created.append(e)
    # Бета — только день 2
    e = TimesheetEntry(
        employee_id=employee_one.id,
        work_date=date(YEAR, MONTH, 2),
        company_id=company_beta.id,
        hours=4,
    )
    db_session.add(e)
    created.append(e)
    db_session.commit()
    return created


@pytest.fixture
def entries_for_emp2(
    db_session: Session,
    employee_two: Employee,
    company_alpha: Company,
) -> list[TimesheetEntry]:
    """Петров: 8ч в Альфа на дни 1, 17."""
    created = []
    for day in [1, 17]:
        e = TimesheetEntry(
            employee_id=employee_two.id,
            work_date=date(YEAR, MONTH, day),
            company_id=company_alpha.id,
            hours=8,
        )
        db_session.add(e)
        created.append(e)
    db_session.commit()
    return created


@pytest.fixture
def calendar_june(db_session: Session) -> ProductionCalendar:
    """Производственный календарь июня 2026 с праздниками."""
    # 12 июня — День России (праздник)
    cal = ProductionCalendar(
        year=YEAR,
        data={
            "months": [
                {
                    "month": MONTH,
                    "days": "7,12,13,14,20,21,27,28",  # выходные + праздник 12
                }
            ]
        },
        source="test",
    )
    db_session.add(cal)
    db_session.commit()
    db_session.refresh(cal)
    return cal


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_workbook(response) -> object:
    """Загружает openpyxl Workbook из ответа."""
    assert response.status_code == 200, response.text
    assert "spreadsheetml" in response.headers["content-type"]
    return load_workbook(BytesIO(response.content))


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_export_returns_valid_xlsx(
    client: TestClient,
    admin_emp: Employee,
    employee_one: Employee,
    entries_for_emp1,
    company_alpha: Company,
    company_beta: Company,
):
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    assert wb is not None
    ws = wb.active
    assert ws is not None


def test_export_header_contains_org_and_t13(
    client: TestClient,
    admin_emp: Employee,
    employee_one: Employee,
    entries_for_emp1,
):
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    ws = wb.active
    # Собираем все строковые значения листа
    all_values = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row
        if cell.value is not None
    )
    assert "ЗЕМЛЯ МО" in all_values
    assert "Т-13" in all_values
    assert str(YEAR) in all_values


def test_export_correct_row_count(
    client: TestClient,
    admin_emp: Employee,
    employee_one: Employee,
    employee_two: Employee,
    entries_for_emp1,
    entries_for_emp2,
    company_alpha: Company,
    company_beta: Company,
):
    """Иванов работал в 2 компаниях → 2 строки. Петров в 1 → 1 строка. Итого 3 data-строки."""
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    ws = wb.active
    # Ищем строки где колонка E (компания, col=5) содержит название компании
    company_names = {"ООО Альфа", "ООО Бета"}
    found = sum(
        1 for row in ws.iter_rows()
        if any(cell.value in company_names for cell in row)
    )
    assert found == 3  # 2 для Иванова + 1 для Петрова


def test_export_hours_match_entries(
    client: TestClient,
    admin_emp: Employee,
    employee_one: Employee,
    entries_for_emp1,
    company_alpha: Company,
):
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    ws = wb.active
    from app.services.timesheet_export import _COL_COMPANY, _total_col

    # Итого Ч компании для Иванова/Альфа = 8*5 = 40 ч (per-row, _total_col)
    total_days = 30  # июнь
    found_40 = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == _COL_COMPANY and cell.value == "ООО Альфа":
                if ws.cell(row=cell.row, column=_total_col(total_days)).value == 40:
                    found_40 = True
    assert found_40, "Не найдено итого 40ч для Иванова/Альфа"


def test_export_empty_period_still_generates(
    client: TestClient,
    admin_emp: Employee,
):
    """Если нет entries вообще — возвращается валидный Excel с пустой таблицей."""
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    all_values = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row
        if cell.value is not None
    )
    assert "ЗЕМЛЯ МО" in all_values  # шапка есть


def test_export_holiday_cell_has_red_fill(
    client: TestClient,
    admin_emp: Employee,
    employee_one: Employee,
    entries_for_emp1,
    calendar_june: ProductionCalendar,
):
    """День 12 июня (День России) должен иметь красный фон в шапке."""
    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    ws = wb.active
    # Ищем ячейку с "12" в шапке (одна из верхних строк, содержит "12")
    red_fill_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and "12" in str(cell.value):
                fill = cell.fill
                if fill and fill.fgColor and fill.fgColor.rgb == "FFFFCCCC":
                    red_fill_found = True
    assert red_fill_found, "Не найдена красная заливка для дня 12"


def test_export_per_company_rows_columns(
    client: TestClient,
    admin_emp: Employee,
    dept: Department,
    company_alpha: Company,
    company_beta: Company,
    db_session: Session,
):
    """Строки по компаниям: Итого Ч компании (per-row), Сверхур./Праздн. Ч,
    общий Итого Ч (merge) считаются корректно."""
    from app.models.schedules import Schedule
    from app.services.timesheet_export import (
        _COL_COMPANY,
        _grand_total_col,
        _norm_col,
        _ot_hours_col,
        _total_col,
    )

    sch = Schedule(name="5/2", hours_per_shift=8, is_active=True, schedule_type="standard")
    db_session.add(sch)
    db_session.commit()
    db_session.refresh(sch)

    emp = Employee(
        full_name="Сверхурочный Работник",
        tab_number="9999",
        department_id=dept.id,
        schedule_id=sch.id,
        is_active=True,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)

    # День 1 (Пн) — 10ч в Альфа; День 3 (Ср) — 8ч в Бета.
    # Переработка теперь помесячная (3.11b п.0): 18ч < месячной нормы → переработки нет.
    db_session.add(TimesheetEntry(employee_id=emp.id, work_date=date(YEAR, MONTH, 1),
                                  company_id=company_alpha.id, hours=10))
    db_session.add(TimesheetEntry(employee_id=emp.id, work_date=date(YEAR, MONTH, 3),
                                  company_id=company_beta.id, hours=8))
    db_session.commit()

    token = get_token(client, "admin@example.com", "admin123")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    wb = _get_workbook(resp)
    ws = wb.active

    # Заголовки новых колонок присутствуют
    header_text = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )
    assert "Сверхур" in header_text
    assert "Праздн" in header_text
    assert "Норма" in header_text
    assert "ООО Альфа" in header_text
    assert "ООО Бета" in header_text

    # Находим первую строку данных сотрудника (по col=5 = "ООО Альфа")
    start_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == _COL_COMPANY and cell.value == "ООО Альфа":
                start_row = cell.row
                break
        if start_row:
            break
    assert start_row is not None

    total_days = 30  # июнь
    # Альфа создана раньше беты → строка Альфа первая, Бета — вторая
    assert ws.cell(row=start_row, column=_total_col(total_days)).value == 10      # Альфа
    assert ws.cell(row=start_row + 1, column=_total_col(total_days)).value == 8   # Бета
    # Общий итог по сотруднику (merge в первой строке) = 18
    assert ws.cell(row=start_row, column=_grand_total_col(total_days)).value == 18
    # Переработка помесячная: суммарно 18ч < нормы → 0 переработки
    ot_alpha = ws.cell(row=start_row, column=_ot_hours_col(total_days)).value or 0
    ot_beta = ws.cell(row=start_row + 1, column=_ot_hours_col(total_days)).value or 0
    assert ot_alpha + ot_beta == 0
    # Норма июня (standard 8ч) — положительное число
    assert (ws.cell(row=start_row, column=_norm_col(total_days)).value or 0) > 0


def test_export_forbidden_for_employee(
    client: TestClient,
    worker_emp: Employee,
):
    token = get_token(client, "worker@example.com", "pass")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 403


def test_export_manager_cannot_see_other_dept(
    client: TestClient,
    manager_emp: Employee,
    dept: Department,
    db_session: Session,
):
    """Manager с dept_id=dept.id не может запросить другой dept_id."""
    other_dept = Department(name="Чужой Отдел", code="ЧО", is_active=True)
    db_session.add(other_dept)
    db_session.commit()
    db_session.refresh(other_dept)

    token = get_token(client, "manager@example.com", "pass")
    resp = client.get(
        f"/api/timesheet/{YEAR}/{MONTH}/export/excel",
        params={"department_id": other_dept.id},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 403
