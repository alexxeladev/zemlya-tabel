"""Тесты импорта сотрудников из Excel (task_employee_import)."""
import datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.core.security import hash_password
from app.models.audit_log import AuditLog
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.schedules import Schedule
from app.services.employee_import import (
    COLUMNS,
    EXAMPLE_MARKER,
    normalize_schedule_key,
    parse_date,
    parse_decimal,
    parse_pay_type,
    parse_weekend_pay_type,
)
from tests.conftest import get_token

# Заведомо валидная строка (компания по коду, окладная)
ROW_OK = ["100", "Сидоров Сидор", "ZMO", "ИТО", "Слесарь", "5/2",
          "окладная", "50 000", "", "коэффициент", "1,5", "01.03.2026"]


@pytest.fixture
def refs(db_session):
    """Справочники: компании, отдел, графики."""
    company = Company(code="ZMO", name='ООО "Комфорт"')
    other = Company(code="KFT", name="Земля МО")
    dept = Department(name="ИТО", code="ITO")
    weekday = Schedule(name="5/2", hours_per_shift=8, schedule_type="weekday")
    cyclic = Schedule(
        name="2/2 смена 1", hours_per_shift=12, schedule_type="cyclic",
        cycle_start_date=datetime.date(2026, 5, 31), cycle_work_days=2, cycle_off_days=2,
    )
    db_session.add_all([company, other, dept, weekday, cyclic])
    db_session.commit()
    for obj in (company, other, dept, weekday, cyclic):
        db_session.refresh(obj)
    return {
        "company": company, "other": other, "dept": dept,
        "weekday": weekday, "cyclic": cyclic,
    }


@pytest.fixture
def admin_token(client, admin_user):
    return get_token(client, "admin@example.com", "admin123")


def auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def make_file(rows: list[list], with_example: bool = True) -> bytes:
    """Собрать .xlsx как заполненный шаблон: заголовки, пример, затем данные."""
    wb = Workbook()
    ws = wb.active
    ws.append([c.title for c in COLUMNS])
    if with_example:
        ws.append([c.example for c in COLUMNS])
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload(client, token: str, content: bytes, confirm: bool = False):
    return client.post(
        "/api/employees/import",
        params={"confirm": confirm},
        files={"file": ("employees.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth(token),
    )


# ── Шаблон ────────────────────────────────────────────────────────────────────

def test_template_download(client, admin_token, refs):
    resp = client.get("/api/employees/import/template", headers=auth(admin_token))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "shablon_sotrudnikov.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    ws = wb["Сотрудники"]
    assert [c.value for c in ws[1]] == [c.title for c in COLUMNS]
    # вторая строка — пример, помечен явно
    assert ws.cell(row=2, column=1).value == EXAMPLE_MARKER
    assert ws.cell(row=2, column=2).value == "Иванов Иван Иванович"

    # лист со справочниками — чтобы заполняющий знал допустимые значения
    ref_ws = wb["Справочники"]
    values = {c.value for row in ref_ws.iter_rows() for c in row if c.value}
    assert 'ZMO — ООО "Комфорт"' in values
    assert "ИТО" in values
    assert "5/2" in values


def test_import_is_admin_only(client, db_session, refs):
    """Импорт — только admin: бухгалтер получает 403, аноним — 401."""
    db_session.add(Employee(
        full_name="Бухгалтер", email="acc@example.com", role="accountant",
        is_active=True, hashed_password=hash_password("acc12345"),
    ))
    db_session.commit()
    token = get_token(client, "acc@example.com", "acc12345")

    assert client.get("/api/employees/import/template", headers=auth(token)).status_code == 403
    assert upload(client, token, make_file([ROW_OK])).status_code == 403
    assert client.get("/api/employees/import/template").status_code == 401


# ── Нормализация значений (чистые функции) ─────────────────────────────────────

@pytest.mark.parametrize("raw,expected", [
    ("5\\2", "5/2"),
    ("5 / 2", "5/2"),
    ("  5/2  ", "5/2"),
    ("5\\2 ", "5/2"),
    ("2/2 Смена 1", "2/2 смена 1"),
])
def test_normalize_schedule_key(raw, expected):
    assert normalize_schedule_key(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    ("50 000", Decimal("50000")),
    ("50 000", Decimal("50000")),        # неразрывный пробел
    ("50 000,50", Decimal("50000.50")),
    ("1.5", Decimal("1.5")),
    ("2500₽", Decimal("2500")),
    (50000, Decimal("50000")),
    (50000.0, Decimal("50000")),
    ("", None),
    (None, None),
])
def test_parse_decimal(raw, expected):
    assert parse_decimal(raw) == expected


def test_parse_decimal_rejects_garbage():
    with pytest.raises(ValueError):
        parse_decimal("пятьдесят тысяч")


@pytest.mark.parametrize("raw,expected", [
    ("01.03.2026", datetime.date(2026, 3, 1)),
    ("2026-03-01", datetime.date(2026, 3, 1)),
    ("01/03/2026", datetime.date(2026, 3, 1)),
    (datetime.datetime(2026, 3, 1, 12, 0), datetime.date(2026, 3, 1)),
    (datetime.date(2026, 3, 1), datetime.date(2026, 3, 1)),
    (46082, datetime.date(2026, 3, 1)),       # серийная дата Excel
    ("", None),
])
def test_parse_date(raw, expected):
    assert parse_date(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    ("окладная", "salary"), ("Оклад", "salary"), ("salary", "salary"), ("", "salary"),
    ("посменная", "per_shift"), ("смена", "per_shift"), ("за смену", "per_shift"),
    ("непонятно", None),
])
def test_parse_pay_type(raw, expected):
    assert parse_pay_type(raw) == expected


@pytest.mark.parametrize("raw,expected", [
    ("коэффициент", "coefficient"), ("коэф", "coefficient"), ("×1.5", "coefficient"),
    ("1,5", "coefficient"), ("", "coefficient"),
    ("фикс", "fixed_rate"), ("фиксированная", "fixed_rate"), ("740", "fixed_rate"),
    ("непонятно", None),
])
def test_parse_weekend_pay_type(raw, expected):
    assert parse_weekend_pay_type(raw) == expected


# ── Превью: распознавание и статусы строк ─────────────────────────────────────

def test_preview_recognizes_values(client, admin_token, refs):
    """Умное распознавание: `5\\2`, компания в кавычках, «50 000», дата ДД.ММ.ГГГГ."""
    content = make_file([
        ["001", "Петров Пётр Петрович", 'ООО "Комфорт"', "ито", "Инженер", "5\\2",
         "окладная", "50 000", "", "коэффициент", "1,5", "01.03.2026"],
    ])
    resp = upload(client, admin_token, content)
    assert resp.status_code == 200
    data = resp.json()

    assert data["confirmed"] is False
    assert (data["total"], data["valid_count"], data["error_count"]) == (1, 1, 0)

    row = data["rows"][0]
    assert row["is_valid"] is True
    assert row["errors"] == []
    assert row["row_number"] == 3          # 1 — заголовки, 2 — пример
    assert row["full_name"] == "Петров Пётр Петрович"
    assert row["company_id"] == refs["company"].id
    assert row["department_id"] == refs["dept"].id
    assert row["schedule_id"] == refs["weekday"].id
    assert Decimal(row["rate"]) == Decimal("50000")
    assert row["pay_type"] == "salary"
    assert row["weekend_pay_type"] == "coefficient"
    assert Decimal(row["weekend_coefficient"]) == Decimal("1.5")
    assert row["hire_date"] == "2026-03-01"


def test_preview_matches_company_by_code_and_short_name(client, admin_token, refs):
    """Компания ищется по коду и по короткому имени без правовой формы."""
    content = make_file([
        [None, "Первый", "ZMO", "", "", "", "", "50000", "", "", "", ""],
        [None, "Второй", "Комфорт", "", "", "", "", "50000", "", "", "", ""],
        [None, "Третий", "  ооо   «комфорт» ", "", "", "", "", "50000", "", "", "", ""],
    ])
    rows = upload(client, admin_token, content).json()["rows"]
    assert [r["company_id"] for r in rows] == [refs["company"].id] * 3
    assert all(r["is_valid"] for r in rows)


def test_preview_example_row_is_skipped(client, admin_token, refs):
    """Строка-пример из шаблона в импорт не попадает."""
    data = upload(client, admin_token, make_file([ROW_OK])).json()
    assert data["total"] == 1
    assert data["rows"][0]["full_name"] == "Сидоров Сидор"


def test_preview_reports_each_error_kind(client, admin_token, refs):
    content = make_file([
        [None, "", "ZMO", "", "", "", "", "50000", "", "", "", ""],
        [None, "Без компании", "", "", "", "", "", "50000", "", "", "", ""],
        [None, "Чужая компания", "ООО Ромашка", "", "", "", "", "50000", "", "", "", ""],
        [None, "Чужой отдел", "ZMO", "Продажи", "", "", "", "50000", "", "", "", ""],
        [None, "Чужой график", "ZMO", "", "", "7/0", "", "50000", "", "", "", ""],
        [None, "Без оклада", "ZMO", "", "", "", "окладная", "", "", "", "", ""],
        [None, "Без ставки", "ZMO", "", "", "", "посменная", "", "", "", "", ""],
        [None, "Оклад не число", "ZMO", "", "", "", "", "много", "", "", "", ""],
        [None, "Дата кривая", "ZMO", "", "", "", "", "50000", "", "", "", "вчера"],
    ])
    data = upload(client, admin_token, content).json()
    assert data["valid_count"] == 0
    assert data["error_count"] == 9

    errors = ["; ".join(r["errors"]) for r in data["rows"]]
    assert "ФИО обязательно" in errors[0]
    assert "Компания обязательна" in errors[1]
    assert "Компания «ООО Ромашка» не найдена" in errors[2]
    assert "Отдел «Продажи» не найден" in errors[3]
    assert "График «7/0» не найден" in errors[4]
    assert "Не указан оклад" in errors[5]
    assert "Не указана ставка за смену" in errors[6]
    assert "Оклад не число" in errors[7]
    assert "не распознана" in errors[8]


def test_preview_duplicate_tab_number(client, db_session, admin_token, refs):
    """Дубль таб.№ — и с существующим в БД, и внутри файла."""
    db_session.add(Employee(full_name="Старый", tab_number="777", is_active=True))
    db_session.commit()

    content = make_file([
        ["777", "Дубль с БД", "ZMO", "", "", "", "", "50000", "", "", "", ""],
        ["555", "Первый", "ZMO", "", "", "", "", "50000", "", "", "", ""],
        ["555", "Дубль в файле", "ZMO", "", "", "", "", "50000", "", "", "", ""],
    ])
    rows = upload(client, admin_token, content).json()["rows"]

    assert rows[0]["errors"] == ["Таб.№ «777» уже существует"]
    assert rows[1]["is_valid"] is True
    assert rows[2]["errors"] == ["Таб.№ «555» уже существует"]


def test_preview_per_shift_row(client, admin_token, refs):
    content = make_file([
        ["S-1", "Сменщик", "ZMO", "ИТО", "Оператор", "2/2 смена 1",
         "посменная", "", "2 500", "фикс", "740", ""],
    ])
    row = upload(client, admin_token, content).json()["rows"][0]
    assert row["is_valid"] is True
    assert row["pay_type"] == "per_shift"
    assert Decimal(row["shift_rate"]) == Decimal("2500")
    assert row["rate"] is None
    assert row["weekend_pay_type"] == "fixed_rate"
    assert Decimal(row["weekend_fixed_rate"]) == Decimal("740")
    assert row["schedule_id"] == refs["cyclic"].id


def test_preview_rejects_broken_file(client, admin_token, refs):
    resp = client.post(
        "/api/employees/import",
        files={"file": ("employees.xlsx", b"not an xlsx at all", "application/octet-stream")},
        headers=auth(admin_token),
    )
    assert resp.status_code == 422
    assert "xlsx" in resp.json()["detail"]


def test_preview_rejects_empty_file(client, admin_token, refs):
    resp = upload(client, admin_token, make_file([]))
    assert resp.status_code == 422
    assert "нет строк" in resp.json()["detail"]


def test_preview_does_not_create_employees(client, db_session, admin_token, refs):
    before = db_session.query(Employee).count()
    upload(client, admin_token, make_file([ROW_OK]))
    assert db_session.query(Employee).count() == before


# ── Подтверждённый импорт ─────────────────────────────────────────────────────

def test_confirmed_import_creates_employee(client, db_session, admin_token, refs):
    data = upload(client, admin_token, make_file([ROW_OK]), confirm=True).json()

    assert data["confirmed"] is True
    assert (data["created_count"], data["skipped_count"]) == (1, 0)
    assert data["rows"][0]["created"] is True

    emp = db_session.query(Employee).filter(Employee.full_name == "Сидоров Сидор").one()
    assert emp.tab_number == "100"
    assert emp.position == "Слесарь"
    assert emp.default_company_id == refs["company"].id
    assert emp.department_id == refs["dept"].id
    assert emp.schedule_id == refs["weekday"].id
    assert emp.pay_type == "salary"
    assert emp.rate == Decimal("50000")
    assert emp.shift_rate is None
    assert emp.weekend_pay_type == "coefficient"
    assert emp.weekend_coefficient == Decimal("1.5")
    assert emp.hire_date == datetime.date(2026, 3, 1)
    assert emp.is_active is True
    # Дефолты карточки, которых нет в шаблоне
    assert emp.holiday_coefficient == Decimal("1.5")
    assert emp.overtime_coefficient == Decimal("1.5")


def test_confirmed_import_does_not_create_access(client, db_session, admin_token, refs):
    """Доступы не импортируются: ни email, ни роли, ни пароля."""
    upload(client, admin_token, make_file([ROW_OK]), confirm=True)
    emp = db_session.query(Employee).filter(Employee.full_name == "Сидоров Сидор").one()
    assert emp.email is None
    assert emp.role is None
    assert emp.hashed_password is None
    assert emp.is_system_admin is False


def test_confirmed_import_per_shift(client, db_session, admin_token, refs):
    content = make_file([
        ["S-1", "Сменщик", "ZMO", "", "Оператор", "2/2 смена 1",
         "посменная", "", "2 500", "фикс", "740", ""],
    ])
    assert upload(client, admin_token, content, confirm=True).json()["created_count"] == 1

    emp = db_session.query(Employee).filter(Employee.full_name == "Сменщик").one()
    assert emp.pay_type == "per_shift"
    assert emp.shift_rate == Decimal("2500")
    assert emp.rate is None
    assert emp.weekend_pay_type == "fixed_rate"
    assert emp.weekend_fixed_rate == Decimal("740")
    assert emp.department_id is None      # пусто → без отдела
    assert emp.schedule_id == refs["cyclic"].id


def test_confirmed_import_partial(client, db_session, admin_token, refs):
    """5 строк: 3 валидных, 1 без компании, 1 дубль таб.№ → создано 3, пропущено 2."""
    db_session.add(Employee(full_name="Старый", tab_number="777", is_active=True))
    db_session.commit()

    content = make_file([
        ["101", "Первый Валидный", "ZMO", "ИТО", "Инженер", "5\\2", "окладная", "50 000", "",
         "коэффициент", "1,5", "01.03.2026"],
        ["102", "Второй Валидный", 'ООО "Комфорт"', "", "", "", "", "60000", "", "", "", ""],
        ["103", "Третий Валидный", "Земля МО", "ИТО", "Оператор", "2/2 смена 1", "посменная",
         "", "2500", "фикс", "740", ""],
        ["104", "Без Компании", "", "ИТО", "", "", "", "50000", "", "", "", ""],
        ["777", "Дубль Табельного", "ZMO", "", "", "", "", "50000", "", "", "", ""],
    ])

    preview = upload(client, admin_token, content).json()
    assert (preview["valid_count"], preview["error_count"]) == (3, 2)

    data = upload(client, admin_token, content, confirm=True).json()
    assert (data["created_count"], data["skipped_count"]) == (3, 2)

    names = {e.full_name for e in db_session.query(Employee).all()}
    assert {"Первый Валидный", "Второй Валидный", "Третий Валидный"} <= names
    assert "Без Компании" not in names
    assert "Дубль Табельного" not in names


def test_confirmed_import_writes_audit_log(client, db_session, admin_token, refs):
    upload(client, admin_token, make_file([ROW_OK]), confirm=True)

    actions = [
        a.action for a in db_session.query(AuditLog).filter(AuditLog.entity_type == "employee")
    ]
    assert "create" in actions
    summary = db_session.query(AuditLog).filter(
        AuditLog.action == "employees_imported"
    ).one()
    assert summary.after == {"created": 1, "skipped": 0, "total": 1}


def test_repeated_confirmed_import_is_blocked_by_tab_number(client, db_session, admin_token, refs):
    """Второй заход тем же файлом ничего не задваивает — таб.№ уже занят."""
    upload(client, admin_token, make_file([ROW_OK]), confirm=True)
    data = upload(client, admin_token, make_file([ROW_OK]), confirm=True).json()

    assert (data["created_count"], data["skipped_count"]) == (0, 1)
    assert "уже существует" in data["rows"][0]["errors"][0]
    assert db_session.query(Employee).filter(Employee.tab_number == "100").count() == 1
