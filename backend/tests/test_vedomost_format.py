"""Порядок юрлиц, их названия и формат Excel-ведомости (task_vedomost_format).

Три части задачи:
1. `companies.sort_order` — настраиваемый порядок, применяется ВЕЗДЕ одной
   функцией сортировки (`app/services/company_order.py`);
2. подписи юрлиц — названиями, а не кодами (`display_name`);
3. Excel-ведомость по образцу финдира: шапка организация/подразделение/месяц,
   полный состав колонок, итоги СВЕРХУ и снизу.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.production_calendars import ProductionCalendar
from app.models.schedules import Schedule
from app.models.timesheet_entries import TimesheetEntry
from app.services.company_order import (
    company_display_name,
    order_index,
    sort_companies,
    sort_company_ids,
)
from tests.conftest import get_token

MAY_BASIC = {"year": 2026, "months": [{"month": 5, "days": "3,4,10,11,17,18,24,25,31"}]}
MAY_WORKDAYS = [d for d in range(1, 32) if d not in (3, 4, 10, 11, 17, 18, 24, 25, 31)]


# ── Фикстуры ──────────────────────────────────────────────────────────────────

@pytest.fixture
def admin(db_session: Session) -> Employee:
    emp = Employee(
        full_name="Админ", email="admin@example.com",
        hashed_password=hash_password("admin123"), role="admin",
        is_active=True, must_change_password=False,
        # Системный пользователь не попадает в табель и ведомость — иначе он
        # встал бы первой строкой и сдвинул проверки на сотрудника.
        is_system_admin=True,
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


@pytest.fixture
def companies(db_session: Session) -> list[Company]:
    """Три юрлица, заведённые в порядке, обратном желаемому."""
    made = []
    for i, (code, name) in enumerate(
        [("ghs", 'ООО "ГХС"'), ("sd", 'ООО "Стройдепартамент"'), ("zmo", 'ООО "ЗМО"')],
        start=1,
    ):
        c = Company(code=code, name=name, is_active=True, sort_order=i)
        db_session.add(c)
        made.append(c)
    db_session.commit()
    for c in made:
        db_session.refresh(c)
    return made


@pytest.fixture
def calendar(db_session: Session) -> ProductionCalendar:
    cal = ProductionCalendar(year=2026, data=MAY_BASIC, source="manual")
    db_session.add(cal)
    db_session.commit()
    return cal


@pytest.fixture
def schedule(db_session: Session) -> Schedule:
    sch = Schedule(
        name="5/2", schedule_type="weekday", hours_per_shift=8, is_active=True,
    )
    db_session.add(sch)
    db_session.commit()
    db_session.refresh(sch)
    return sch


@pytest.fixture
def department(db_session: Session, companies: list[Company]) -> Department:
    """Отдел с головной компанией — она попадёт в шапку выгрузки."""
    zmo = next(c for c in companies if c.code == "zmo")
    dept = Department(
        name="Секьюрити", code="SEC", is_active=True, head_company_id=zmo.id,
    )
    db_session.add(dept)
    db_session.commit()
    db_session.refresh(dept)
    return dept


@pytest.fixture
def worker(
    db_session: Session, department: Department, schedule: Schedule,
    companies: list[Company],
) -> Employee:
    zmo = next(c for c in companies if c.code == "zmo")
    emp = Employee(
        full_name="Стрателюк Евгений", tab_number="0000-00235",
        position="Руководитель ОП", is_active=True, must_change_password=False,
        department_id=department.id, schedule_id=schedule.id,
        default_company_id=zmo.id, rate=Decimal("150000"),
    )
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    for day in MAY_WORKDAYS:
        db_session.add(TimesheetEntry(
            employee_id=emp.id, position_id=emp.primary_position.id,
            work_date=date(2026, 5, day), company_id=zmo.id, hours=8,
        ))
    db_session.commit()
    return emp


# ── ЧАСТЬ 1: единая сортировка ────────────────────────────────────────────────

class TestCompanyOrder:
    def test_sort_companies_uses_sort_order_not_id(self, db_session, companies):
        """Порядок задаётся sort_order, а не порядком заведения."""
        zmo = next(c for c in companies if c.code == "zmo")
        ghs = next(c for c in companies if c.code == "ghs")
        zmo.sort_order, ghs.sort_order = 1, 9
        db_session.commit()
        assert [c.code for c in sort_companies(companies)] == ["zmo", "sd", "ghs"]

    def test_equal_sort_order_falls_back_to_id(self, db_session, companies):
        """Дубль sort_order — штатное состояние после ручной правки; порядок
        всё равно детерминированный, второй ключ — id."""
        for c in companies:
            c.sort_order = 5
        db_session.commit()
        by_id = sorted(companies, key=lambda c: c.id)
        assert sort_companies(companies) == by_id

    def test_sort_company_ids_puts_unknown_and_none_last(self, companies):
        """Неактивное/неизвестное юрлицо и «нет компании» не теряются, а уходят
        в конец: строку с их часами выкидывать нельзя."""
        by_id = {c.id: c for c in companies}
        ids = [None, 999, companies[2].id, companies[0].id]
        assert sort_company_ids(ids, by_id) == [
            companies[0].id, companies[2].id, 999, None,
        ]

    def test_order_index_fixes_positions(self, companies):
        idx = order_index(c.id for c in sort_companies(companies))
        assert idx[companies[0].id] == 0

    def test_companies_endpoint_returns_configured_order(
        self, client: TestClient, admin, companies, db_session
    ):
        zmo = next(c for c in companies if c.code == "zmo")
        zmo.sort_order = 0
        db_session.commit()
        token = get_token(client, "admin@example.com", "admin123")
        resp = client.get("/api/companies", headers={"Authorization": f"Bearer {token}"})
        assert resp.status_code == 200
        assert [c["code"] for c in resp.json()][0] == "zmo"

    def test_reorder_endpoint_renumbers_densely(
        self, client: TestClient, admin, companies
    ):
        """Перестановка присылает ПОЛНЫЙ порядок и раскладывает 1..N — иначе
        остались бы дыры и совпадающие значения."""
        token = get_token(client, "admin@example.com", "admin123")
        ids = [companies[2].id, companies[0].id, companies[1].id]
        resp = client.put(
            "/api/companies/order", json={"company_ids": ids},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert [c["id"] for c in body] == ids
        assert [c["sort_order"] for c in body] == [1, 2, 3]

    def test_reorder_puts_missing_companies_at_the_end(
        self, client: TestClient, admin, companies
    ):
        """Компания, которой нет в присланном списке, не пропадает."""
        token = get_token(client, "admin@example.com", "admin123")
        resp = client.put(
            "/api/companies/order", json={"company_ids": [companies[2].id]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 200
        assert len(resp.json()) == 3
        assert resp.json()[0]["id"] == companies[2].id

    def test_reorder_rejects_unknown_company(self, client: TestClient, admin, companies):
        token = get_token(client, "admin@example.com", "admin123")
        resp = client.put(
            "/api/companies/order", json={"company_ids": [999999]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 404

    def test_reorder_is_admin_only(
        self, client: TestClient, admin, companies, db_session
    ):
        db_session.add(Employee(
            full_name="Бух", email="acc@example.com",
            hashed_password=hash_password("acc123"), role="accountant",
            is_active=True, must_change_password=False,
        ))
        db_session.commit()
        token = get_token(client, "acc@example.com", "acc123")
        resp = client.put(
            "/api/companies/order", json={"company_ids": [companies[0].id]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 403

    def test_manual_sort_order_edit_renumbers(
        self, client: TestClient, admin, companies, db_session
    ):
        """Ручной ввод числа может дать дубль; после правки значения
        нормализуются в плотный 1..N."""
        token = get_token(client, "admin@example.com", "admin123")
        # Ставим 0, а не 1: при равном sort_order порядок разводит id, и
        # «1» встала бы ПОСЛЕ уже существующей единицы.
        resp = client.patch(
            f"/api/companies/{companies[2].id}", json={"sort_order": 0},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 200
        listed = client.get(
            "/api/companies", headers={"Authorization": f"Bearer {token}"}
        ).json()
        assert [c["sort_order"] for c in listed] == [1, 2, 3]
        assert listed[0]["id"] == companies[2].id

    def test_new_company_goes_to_the_end(self, client: TestClient, admin, companies):
        """Новая компания не вклинивается в настроенный порядок."""
        token = get_token(client, "admin@example.com", "admin123")
        resp = client.post(
            "/api/companies", json={"code": "new", "name": 'ООО "Новое"'},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 201
        listed = client.get(
            "/api/companies", headers={"Authorization": f"Bearer {token}"}
        ).json()
        assert listed[-1]["code"] == "new"

    def test_statement_columns_follow_configured_order(
        self, client: TestClient, admin, companies, calendar, worker, db_session
    ):
        """Ведомость (а с ней и Excel) перечисляет юрлица тем же порядком."""
        token = get_token(client, "admin@example.com", "admin123")
        client.put(
            "/api/companies/order",
            json={"company_ids": [c.id for c in reversed(companies)]},
            headers={"Authorization": f"Bearer {token}"},
        )
        resp = client.get(
            "/api/timesheet/2026/5/statement",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 200
        assert [c["code"] for c in resp.json()["companies"]] == ["zmo", "sd", "ghs"]

    def test_org_tree_follows_configured_order(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Дерево «Оргструктуры» — то место, где порядок и правится стрелками.

        Регрессия: build_org_tree пересортировывал компании ПО ИМЕНИ поверх
        ORDER BY запроса, и стрелки ▲▼ не двигали узлы вовсе.
        """
        token = get_token(client, "admin@example.com", "admin123")
        client.put(
            "/api/companies/order",
            json={"company_ids": [c.id for c in reversed(companies)]},
            headers={"Authorization": f"Bearer {token}"},
        )
        resp = client.get("/api/org/tree", headers={"Authorization": f"Bearer {token}"})
        assert resp.status_code == 200
        assert [c["code"] for c in resp.json()["companies"]] == ["zmo", "sd", "ghs"]

    def test_import_template_reference_sheet_follows_configured_order(
        self, client: TestClient, admin, companies
    ):
        """Список юрлиц в шаблоне импорта — тот же порядок (он там подсказка,
        по которой заполняют файл)."""
        from io import BytesIO

        import openpyxl

        token = get_token(client, "admin@example.com", "admin123")
        client.put(
            "/api/companies/order",
            json={"company_ids": [c.id for c in reversed(companies)]},
            headers={"Authorization": f"Bearer {token}"},
        )
        resp = client.get(
            "/api/employees/import/template",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 200
        ws = openpyxl.load_workbook(BytesIO(resp.content))["Справочники"]
        # Юрлица перечислены как «код — название»: код здесь не украшение, а
        # принимаемый парсером ключ, поэтому в этом листе он остаётся.
        listed = [
            c.value for col in ws.iter_cols(values_only=False) for c in col
            if isinstance(c.value, str) and " — " in c.value and c.value.endswith('"')
        ]
        assert listed == [
            'zmo — ООО "ЗМО"', 'sd — ООО "Стройдепартамент"', 'ghs — ООО "ГХС"',
        ]

    def test_timesheet_companies_follow_configured_order(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        token = get_token(client, "admin@example.com", "admin123")
        client.put(
            "/api/companies/order",
            json={"company_ids": [c.id for c in reversed(companies)]},
            headers={"Authorization": f"Bearer {token}"},
        )
        resp = client.get(
            "/api/timesheet/2026/5", headers={"Authorization": f"Bearer {token}"}
        )
        assert resp.status_code == 200
        assert [c["code"] for c in resp.json()["companies"]] == ["zmo", "sd", "ghs"]


# ── ЧАСТЬ 2: названия вместо кодов ────────────────────────────────────────────

class TestCompanyDisplayName:
    def test_strips_legal_form_and_quotes(self, db_session):
        c = Company(code="kft", name='ООО "Комфорт"', is_active=True)
        assert company_display_name(c) == "Комфорт"

    def test_manual_short_name_wins(self, db_session):
        c = Company(
            code="exp", name='ООО "Комфорт-Эксплуатация"',
            short_name="К-Эксплуат.", is_active=True,
        )
        assert company_display_name(c) == "К-Эксплуат."

    def test_falls_back_to_code_when_no_name(self):
        assert company_display_name(Company(code="zmo", name="")) == "zmo"

    def test_read_schema_exposes_display_name(
        self, client: TestClient, admin, companies
    ):
        token = get_token(client, "admin@example.com", "admin123")
        listed = client.get(
            "/api/companies", headers={"Authorization": f"Bearer {token}"}
        ).json()
        assert {c["display_name"] for c in listed} == {"ЗМО", "Стройдепартамент", "ГХС"}

    def test_statement_company_ref_carries_display_name(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        token = get_token(client, "admin@example.com", "admin123")
        resp = client.get(
            "/api/timesheet/2026/5/statement",
            headers={"Authorization": f"Bearer {token}"},
        )
        refs = {c["code"]: c for c in resp.json()["companies"]}
        assert refs["zmo"]["display_name"] == "ЗМО"
        # Полное имя остаётся — оно уходит в подсказку
        assert refs["zmo"]["name"] == 'ООО "ЗМО"'


# ── ЧАСТЬ 3: Excel-ведомость по образцу ───────────────────────────────────────

def _statement_sheet(client: TestClient, token: str, **params):
    resp = client.get(
        "/api/timesheet/2026/5/statement/export/excel",
        params=params, headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    return wb.active


# Заголовки колонок 1..27, в порядке образца.
EXPECTED_FIXED = [
    "№ п/п", "табельный номер", "ФИО", "Компания", "Подразделение", "Должность",
    "ОКЛАД", "график работы", "Кол-во дней отпуска", "Кол-во дней больничного",
    "Норма дней", "Факт дней", "НОРМА Кол-во раб. час в месяце",
    "ФАКТ Кол-во отраб. час в месяце", "Учет переработок 0; 1; 1,5",
    "Кол-во переработки, час", "Сумма ПЕРЕРАБОТКи", "Начислено, оклад",
    "Выплачено отпуск/больничный", "Премия Базовая", "KPI по выполнению плана",
    "Премия", "Основание", "Итого начислено", "выплачено аванс\\удержано",
    "Сумма к выплате", "Разбивка по %",
]


class TestStatementExcel:
    def test_heading_names_organization_department_and_month(
        self, client: TestClient, admin, companies, calendar, worker, department
    ):
        """Шапка образца: юрлицо, подразделение, «ВЕДОМОСТЬ … за <Месяц Год>»."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token, department_id=department.id)
        assert ws.cell(1, 2).value == 'ООО "ЗМО"'
        assert ws.cell(2, 2).value == "(наименование организации)"
        assert ws.cell(3, 2).value == '"Секьюрити"'
        assert ws.cell(4, 2).value == "(наименование структурного подразделения)"
        assert ws.cell(5, 2).value == "ВЕДОМОСТЬ на выплату заработной платы за Май 2026"

    def test_heading_falls_back_when_exported_for_all_departments(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Однозначного юрлица у выгрузки «по всем» нет — подписываем группой."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        assert ws.cell(1, 2).value == "ДЕВЕЛОПМЕНТ ГРУППА «ЗЕМЛЯ МО»"
        assert ws.cell(3, 2).value == '"Все подразделения"'

    def test_heading_rows_are_merged_across_the_whole_table(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Merge по ширине таблицы — она зависит от числа юрлиц."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        merged = {str(m) for m in ws.merged_cells.ranges}
        last_col = max(c.column for c in ws[7] if c.value)
        from openpyxl.utils import get_column_letter
        assert f"B1:{get_column_letter(last_col)}1" in merged
        assert f"B5:{get_column_letter(last_col)}5" in merged

    def test_fixed_columns_match_the_template_order(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        got = [ws.cell(7, 2 + i).value for i in range(len(EXPECTED_FIXED))]
        assert got == EXPECTED_FIXED

    def test_company_columns_come_after_percent_split_by_name_and_order(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Колонки юрлиц — названиями и в настроенном порядке (ч.1 + ч.2)."""
        token = get_token(client, "admin@example.com", "admin123")
        client.put(
            "/api/companies/order",
            json={"company_ids": [c.id for c in reversed(companies)]},
            headers={"Authorization": f"Bearer {token}"},
        )
        ws = _statement_sheet(client, token)
        first = 2 + len(EXPECTED_FIXED)
        assert [ws.cell(7, first + i).value for i in range(3)] == [
            "ЗМО", "Стройдепартамент", "ГХС",
        ]
        assert ws.cell(7, first + 3).value == "ИТОГО Разбивка"
        assert ws.cell(7, first + 4).value == "ПРИМЕЧАНИЕ"

    def test_reason_and_night_columns_stay_in_the_tail(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Того, чего нет в образце, — только ХВОСТОМ: в шаблоне финдира буквы
        колонок зафиксированы."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        tail_start = 2 + len(EXPECTED_FIXED) + len(companies) + 2
        assert [ws.cell(7, tail_start + i).value for i in range(5)] == [
            "Обоснование премии", "Обоснование KPI", "Обоснование удержаний",
            "Ночных смен", "Надбавка за ночные",
        ]

    def test_totals_row_sits_above_the_employees(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Необычно, но так в образце: итоги сразу под шапкой."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        accrued_col = 2 + EXPECTED_FIXED.index("Итого начислено")
        # Строка 8 — итоги, строка 9 — первый сотрудник
        assert ws.cell(8, accrued_col).value == 150000
        assert ws.cell(9, 4).value == "Стрателюк Евгений"

    def test_bottom_totals_repeat_the_top_ones(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        bottom = ws.max_row
        assert ws.cell(bottom, 3).value == "ИТОГО"
        for header in ("Итого начислено", "Сумма к выплате", "Начислено, оклад"):
            col = 2 + EXPECTED_FIXED.index(header)
            assert ws.cell(bottom, col).value == ws.cell(8, col).value

    def test_totals_match_the_statement_api(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Excel и веб-ведомость обязаны показывать одни и те же итоги —
        суммы берутся из тех же total_* и здесь не пересчитываются."""
        token = get_token(client, "admin@example.com", "admin123")
        api = client.get(
            "/api/timesheet/2026/5/statement",
            headers={"Authorization": f"Bearer {token}"},
        ).json()
        ws = _statement_sheet(client, token)
        pairs = {
            "Итого начислено": "total_accrued",
            "Сумма к выплате": "total_net_payout",
            "Начислено, оклад": "total_base_salary",
            "Сумма ПЕРЕРАБОТКи": "total_overtime_amount",
        }
        for header, field in pairs.items():
            col = 2 + EXPECTED_FIXED.index(header)
            assert Decimal(str(ws.cell(8, col).value)) == Decimal(api[field])

    def test_row_carries_absence_days_norm_and_fact_days(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """Поля, которых в Excel не было, а в веб-ведомости есть."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        for header, expected in (
            ("Кол-во дней отпуска", 0),
            ("Кол-во дней больничного", 0),
            ("Норма дней", len(MAY_WORKDAYS)),
            ("Факт дней", len(MAY_WORKDAYS)),
            ("НОРМА Кол-во раб. час в месяце", len(MAY_WORKDAYS) * 8),
            ("ФАКТ Кол-во отраб. час в месяце", len(MAY_WORKDAYS) * 8),
        ):
            col = 2 + EXPECTED_FIXED.index(header)
            assert ws.cell(9, col).value == expected, header

    def test_percent_split_is_written_as_text(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """«Разбивка по %» — как в образце: «ЗМО 100%», названием, не кодом."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        col = 2 + EXPECTED_FIXED.index("Разбивка по %")
        assert ws.cell(9, col).value == "ЗМО 100%"

    def test_zero_company_columns_are_zero_not_blank(
        self, client: TestClient, admin, companies, calendar, worker
    ):
        """В образце юрлица без доли стоят нулями — по строке видно, что они
        в распределение не вошли."""
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        first = 2 + len(EXPECTED_FIXED)
        values = [ws.cell(9, first + i).value for i in range(len(companies))]
        assert 0 in values and None not in values

    def test_numbering_is_by_people(
        self, client: TestClient, admin, companies, calendar, worker, db_session,
        department, schedule
    ):
        """№ п/п считается по людям: у совместителя строк несколько, номер один."""
        zmo = next(c for c in companies if c.code == "zmo")
        second = Employee(
            full_name="Второй Сотрудник", tab_number="T-2", is_active=True,
            must_change_password=False, department_id=department.id,
            schedule_id=schedule.id, default_company_id=zmo.id, rate=Decimal("50000"),
        )
        db_session.add(second)
        db_session.commit()
        db_session.add(TimesheetEntry(
            employee_id=second.id, position_id=second.primary_position.id,
            work_date=date(2026, 5, MAY_WORKDAYS[0]), company_id=zmo.id, hours=8,
        ))
        db_session.commit()
        token = get_token(client, "admin@example.com", "admin123")
        ws = _statement_sheet(client, token)
        numbers = [ws.cell(r, 2).value for r in (9, 10)]
        assert numbers == [1, 2]

    def test_export_stays_forbidden_for_timekeeper(
        self, client: TestClient, admin, companies, calendar, worker, db_session
    ):
        """Ведомость — финансы; табельщик её не выгружает (роль не тронута)."""
        db_session.add(Employee(
            full_name="Табельщик", email="tk@example.com",
            hashed_password=hash_password("tk123456"), role="timekeeper",
            is_active=True, must_change_password=False,
        ))
        db_session.commit()
        token = get_token(client, "tk@example.com", "tk123456")
        resp = client.get(
            "/api/timesheet/2026/5/statement/export/excel",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert resp.status_code == 403
