"""База распределения затрат = «Итого начислено» (task_distribution_base_fix).

Откат ошибки task_it_arm_distribution ч.2, где базой сделали «К выплате».
Распределение по юрлицам отражает ЗАТРАТЫ компании на сотрудника, а затраты
возникают в момент НАЧИСЛЕНИЯ. Удержания (займ, аванс) — возврат ранее выданных
средств, отдельная операция: затраты они не уменьшают. Округление «К выплате»
до тысячи на распределение не влияет вовсе.

Суммы по юрлицам остаются кратными 1000 ₽, но получаются ТОЛЬКО floor-ом плюс
раздачей недостающих тысяч по наибольшим хвостам: математическое округление
приписало бы юрлицам больше затрат, чем начислено. Разница «начислено − Σ
распределения» (0…999 ₽) — нераспределённый остаток, он не приписывается никому
и показывается отдельным показателем.

Здесь проверяются требования задачи целиком; регрессия механизмов (заявки/АРМ,
приоритет карточки, целевые премии) — в своих файлах.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.models.companies import Company
from app.models.company_shares import EmployeeCompanyShare
from app.models.departments import Department
from app.models.employee_adjustments import EmployeeAdjustment
from app.models.employees import Employee
from app.models.production_calendars import ProductionCalendar
from app.models.schedules import Schedule
from app.models.timesheet_entries import TimesheetEntry
from app.services.distribution import distribute_largest_remainder
from app.services.payroll_statement import unallocated_remainder
from tests.conftest import get_token

MAY_BASIC = {"year": 2026, "months": [{"month": 5, "days": "3,4,10,11,17,18,24,25,31"}]}
MAY_WORKDAYS = [d for d in range(1, 32) if d not in (3, 4, 10, 11, 17, 18, 24, 25, 31)]
_ZERO = Decimal("0")


# ── Фикстуры ──────────────────────────────────────────────────────────────────

@pytest.fixture
def companies(db_session: Session) -> list[Company]:
    cs = [
        Company(code="BFA", name="База Один", is_active=True, sort_order=1),
        Company(code="BFB", name="База Два", is_active=True, sort_order=2),
    ]
    db_session.add_all(cs)
    db_session.commit()
    for c in cs:
        db_session.refresh(c)
    return cs


@pytest.fixture
def dept(db_session: Session) -> Department:
    d = Department(name="Отдел базы", code="BFD", is_active=True)
    db_session.add(d)
    db_session.commit()
    db_session.refresh(d)
    return d


@pytest.fixture
def schedule(db_session: Session) -> Schedule:
    s = Schedule(name="5/2", hours_per_shift=8, schedule_type="weekday", is_active=True)
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture
def calendar(db_session: Session) -> ProductionCalendar:
    cal = ProductionCalendar(year=2026, data=MAY_BASIC, source="manual")
    db_session.add(cal)
    db_session.commit()
    return cal


@pytest.fixture
def admin(db_session: Session) -> Employee:
    emp = Employee(full_name="Админ базы", email="basefix@example.com",
                   hashed_password=hash_password("admin123"), role="admin",
                   is_active=True, must_change_password=False, is_system_admin=True)
    db_session.add(emp)
    db_session.commit()
    db_session.refresh(emp)
    return emp


def _worker(db: Session, name, tab, rate, dept, company, schedule) -> Employee:
    emp = Employee(full_name=name, tab_number=tab, is_active=True,
                   rate=Decimal(rate), schedule_id=schedule.id,
                   default_company_id=company.id, department_id=dept.id)
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


def _full_norm(db: Session, emp: Employee, company: Company):
    """Полная норма мая на одном юрлице → «Итого начислено» = оклад ровно."""
    for d in MAY_WORKDAYS:
        db.add(TimesheetEntry(employee_id=emp.id, position_id=emp.primary_position.id,
                              work_date=date(2026, 5, d), company_id=company.id, hours=8))
    db.commit()


def _shares(db: Session, emp: Employee, pairs: list[tuple[Company, str]]):
    for company, percent in pairs:
        db.add(EmployeeCompanyShare(
            employee_id=emp.id, position_id=emp.primary_position.id,
            company_id=company.id, percent=Decimal(percent)))
    db.commit()


def _h(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _token(client: TestClient) -> str:
    return get_token(client, "basefix@example.com", "admin123")


def _statement(client: TestClient, token: str) -> dict:
    r = client.get("/api/timesheet/2026/5/statement", headers=_h(token))
    assert r.status_code == 200, r.text
    return r.json()


def _row(statement: dict, employee_id: int) -> dict:
    return next(r for r in statement["rows"] if r["employee_id"] == employee_id)


def _amounts(row: dict) -> dict[int, Decimal]:
    return {d["company_id"]: Decimal(d["amount"]) for d in row["distribution"]}


# ── Ч.1: база = «Итого начислено» ─────────────────────────────────────────────

class TestBaseIsAccrued:
    def test_loan_deduction_does_not_shrink_distribution(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """ОБЯЗАТЕЛЬНЫЙ тест задачи: начислено 100 000, удержан заём 20 000 →
        распределяется 100 000, а не 80 000.

        Компания потратила на сотрудника ровно то, что начислила; погашение
        займа — возврат ранее выданных денег, к затратам на оплату труда
        отношения не имеет.
        """
        emp = _worker(db_session, "Заёмщик", "BF-1", "100000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        emp.loan_amount = Decimal("20000")
        emp.loan_term_months = 1
        emp.loan_start_date = date(2026, 5, 1)
        db_session.commit()
        _shares(db_session, emp, [(companies[0], "60"), (companies[1], "40")])

        row = _row(_statement(client, _token(client)), emp.id)
        assert Decimal(row["accrued_total"]) == Decimal("100000")
        assert Decimal(row["deductions"]) == Decimal("20000")
        assert Decimal(row["net_payout"]) == Decimal("80000")

        amounts = _amounts(row)
        assert sum(amounts.values()) == Decimal("100000")
        assert amounts[companies[0].id] == Decimal("60000")
        assert amounts[companies[1].id] == Decimal("40000")
        assert Decimal(row["unallocated_remainder"]) == _ZERO

    def test_advance_deduction_does_not_shrink_distribution(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Аванс выплачен раньше, но затраты компании полные — 100 000."""
        emp = _worker(db_session, "Авансовый", "BF-2", "100000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add(EmployeeAdjustment(
            employee_id=emp.id, position_id=emp.primary_position.id,
            year=2026, month=5, kind="advance", amount=Decimal("25000"),
            reason="аванс за май"))
        db_session.commit()
        _shares(db_session, emp, [(companies[0], "60"), (companies[1], "40")])

        row = _row(_statement(client, _token(client)), emp.id)
        assert Decimal(row["net_payout"]) == Decimal("75000")
        assert sum(_amounts(row).values()) == Decimal("100000")

    def test_totals_use_accrued_too(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Итоги ведомости считаются от того же начисленного: Σ по юрлицам +
        Σ остатков = Σ «Итого начислено», а не Σ «К выплате»."""
        a = _worker(db_session, "Первый", "BF-3", "100000", dept, companies[0], schedule)
        b = _worker(db_session, "Второй", "BF-4", "70000", dept, companies[1], schedule)
        _full_norm(db_session, a, companies[0])
        _full_norm(db_session, b, companies[1])
        db_session.add(EmployeeAdjustment(
            employee_id=a.id, position_id=a.primary_position.id,
            year=2026, month=5, kind="advance", amount=Decimal("40000"),
            reason="аванс"))
        db_session.commit()

        statement = _statement(client, _token(client))
        totals = {int(k): Decimal(v) for k, v in statement["distribution_totals"].items()}
        accrued = Decimal(statement["total_accrued"])
        rest = Decimal(statement["total_unallocated_remainder"])
        assert accrued == Decimal("170000")
        assert sum(totals.values()) + rest == accrued
        assert sum(totals.values()) != Decimal(statement["total_net_payout"])


# ── Ч.2: только floor, переразнесения нет ─────────────────────────────────────

class TestFloorOnly:
    def test_never_exceeds_base(self):
        """Σ распределения НИКОГДА не больше базы — на любых весах и суммах."""
        for weights in ({1: Decimal(1)},
                        {1: Decimal(3), 2: Decimal(2)},
                        {1: Decimal(45), 2: Decimal(6), 3: Decimal(13), 4: Decimal(7)}):
            for base in ("1", "999", "1000", "1001", "152381", "152700", "99999"):
                total = Decimal(base)
                amounts = distribute_largest_remainder(total, weights)
                assert sum(amounts.values()) <= total, (base, weights)

    def test_152700_is_not_rounded_up_to_153000(self):
        """Случай из задачи: 152 700 математически округлилось бы до 153 000 —
        юрлицам приписали бы на 300 ₽ больше, чем начислено. Только floor."""
        amounts = distribute_largest_remainder(Decimal("152700"), {1: Decimal(1)})
        assert sum(amounts.values()) == Decimal("152000")
        assert unallocated_remainder(Decimal("152700"), amounts) == Decimal("700")

    def test_all_amounts_are_multiples_of_thousand(self):
        weights = {1: Decimal(45), 2: Decimal(6), 3: Decimal(13), 4: Decimal(7)}
        for base in ("57000", "152381", "1000000", "83917"):
            amounts = distribute_largest_remainder(Decimal(base), weights)
            assert all(a % 1000 == 0 for a in amounts.values()), base

    def test_remainder_is_between_zero_and_999(self):
        weights = {1: Decimal(3), 2: Decimal(2), 3: Decimal(1)}
        for base in range(100000, 100010):
            total = Decimal(base)
            rest = unallocated_remainder(
                total, distribute_largest_remainder(total, weights)
            )
            assert _ZERO <= rest <= Decimal("999"), base
            assert rest == total % 1000

    def test_remainder_visible_in_statement_row(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Пример задачи в ведомости: начислено 152 381 → разнесено 152 000,
        остаток 381, все суммы по юрлицам кратны тысяче."""
        emp = _worker(db_session, "Остаточный", "BF-5", "150000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add(EmployeeAdjustment(
            employee_id=emp.id, position_id=emp.primary_position.id,
            year=2026, month=5, kind="premium", amount=Decimal("2381"),
            reason="премия"))
        db_session.commit()
        _shares(db_session, emp, [(companies[0], "60"), (companies[1], "40")])

        row = _row(_statement(client, _token(client)), emp.id)
        assert Decimal(row["accrued_total"]) == Decimal("152381")
        amounts = _amounts(row)
        assert sum(amounts.values()) == Decimal("152000")
        assert Decimal(row["unallocated_remainder"]) == Decimal("381")
        assert Decimal(row["distribution_total"]) == Decimal("152000")
        assert all(a % 1000 == 0 for a in amounts.values())


# ── Ч.3: два РАЗНЫХ показателя остатка ────────────────────────────────────────

class TestTwoIndicators:
    def test_both_indicators_at_once_and_differ(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """«Эффект округления» и «Нераспределённый остаток» — разные формулы и
        разные числа; оба корректны одновременно.

        Начислено 152 381, удержан аванс 2 000 → к выплате точно 150 381,
        округлённо 150 000, хвост +381 (осело в пользу компании). Распределение
        при этом делит 152 381 → 152 000 и оставляет свой остаток 381.
        Совпадение величины здесь случайно — важно, что это разные показатели с
        разными базами: один считается от выплаты, другой от начисления.
        """
        emp = _worker(db_session, "Двойной", "BF-6", "150000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add_all([
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="premium", amount=Decimal("2381"),
                reason="премия"),
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="advance", amount=Decimal("2000"),
                reason="аванс"),
        ])
        db_session.commit()

        row = _row(_statement(client, _token(client)), emp.id)
        # Показатель 1 — округление ВЫПЛАТЫ (знак любой).
        assert Decimal(row["net_payout_exact"]) == Decimal("150381")
        assert Decimal(row["net_payout"]) == Decimal("150000")
        assert Decimal(row["rounding_tail"]) == Decimal("381")
        # Показатель 2 — округление РАСПРЕДЕЛЕНИЯ (всегда ≥ 0), от начисленного.
        assert Decimal(row["accrued_total"]) == Decimal("152381")
        assert sum(_amounts(row).values()) == Decimal("152000")
        assert Decimal(row["unallocated_remainder"]) == Decimal("381")

    def test_indicators_have_different_values_when_deductions_are_odd(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Чтобы совпадение чисел выше не читалось как «один показатель»:
        при другом удержании они расходятся."""
        emp = _worker(db_session, "Разный", "BF-7", "150000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add_all([
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="premium", amount=Decimal("2381"),
                reason="премия"),
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="advance", amount=Decimal("2100"),
                reason="аванс"),
        ])
        db_session.commit()

        row = _row(_statement(client, _token(client)), emp.id)
        assert Decimal(row["net_payout_exact"]) == Decimal("150281")
        assert Decimal(row["rounding_tail"]) == Decimal("281")
        assert Decimal(row["unallocated_remainder"]) == Decimal("381")
        assert Decimal(row["rounding_tail"]) != Decimal(row["unallocated_remainder"])

    def test_dashboard_shows_both(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Дашборд отдаёт оба показателя за период, и они не подменяют друг
        друга: rounding_effect от выплаты, unallocated_remainder от начисления."""
        emp = _worker(db_session, "Дашборд", "BF-8", "150000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add_all([
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="premium", amount=Decimal("2381"),
                reason="премия"),
            EmployeeAdjustment(
                employee_id=emp.id, position_id=emp.primary_position.id,
                year=2026, month=5, kind="advance", amount=Decimal("2100"),
                reason="аванс"),
        ])
        db_session.commit()
        token = _token(client)

        r = client.get("/api/dashboard/2026/5", headers=_h(token))
        assert r.status_code == 200, r.text
        payroll = r.json()["payroll"]
        assert Decimal(payroll["rounding_effect"]) == Decimal("281")
        assert Decimal(payroll["unallocated_remainder"]) == Decimal("381")

        # Цифры дашборда обязаны сходиться с ведомостью — тот же путь расчёта.
        statement = _statement(client, token)
        assert Decimal(payroll["rounding_effect"]) == Decimal(
            statement["total_rounding_tail"]
        )
        assert Decimal(payroll["unallocated_remainder"]) == Decimal(
            statement["total_unallocated_remainder"]
        )

    def test_dashboard_hides_both_from_timekeeper(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Оба показателя — деньги: табельщику блок ФОТ не отдаётся вовсе."""
        keeper = Employee(
            full_name="Табельщик", email="basefixkeeper@example.com",
            hashed_password=hash_password("keep123"), role="timekeeper",
            is_active=True, must_change_password=False, department_id=dept.id)
        keeper.managed_departments = [dept]
        db_session.add(keeper)
        db_session.commit()

        token = get_token(client, "basefixkeeper@example.com", "keep123")
        r = client.get("/api/dashboard/2026/5", headers=_h(token))
        assert r.status_code == 200, r.text
        assert r.json()["payroll"] is None


# ── Выгрузка ──────────────────────────────────────────────────────────────────

class TestExcel:
    def test_excel_carries_remainder_column(
        self, client, db_session, admin, calendar, dept, companies, schedule
    ):
        """Выгрузка самодостаточна: без колонки остатка «ИТОГО Разбивка» меньше
        «Итого начислено» и объяснить это было бы нечем."""
        emp = _worker(db_session, "Экспортный", "BF-9", "150000", dept,
                      companies[0], schedule)
        _full_norm(db_session, emp, companies[0])
        db_session.add(EmployeeAdjustment(
            employee_id=emp.id, position_id=emp.primary_position.id,
            year=2026, month=5, kind="premium", amount=Decimal("2381"),
            reason="премия"))
        db_session.commit()
        token = _token(client)

        r = client.get(
            "/api/timesheet/2026/5/statement/export/excel", headers=_h(token)
        )
        assert r.status_code == 200, r.text
        ws = load_workbook(BytesIO(r.content)).active
        header_row, totals_top_row, first_data_row = 7, 8, 9
        col = next(
            i for i, c in enumerate(ws[header_row], start=1)
            if c.value == "Нераспределённый остаток"
        )
        assert ws.cell(row=first_data_row, column=col).value == 381
        assert ws.cell(row=totals_top_row, column=col).value == 381
