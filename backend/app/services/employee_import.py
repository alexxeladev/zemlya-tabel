"""
Импорт сотрудников из Excel (task_employee_import).

Три вещи в одном месте:
  1. `generate_import_template` — скачиваемый шаблон .xlsx: строка заголовков,
     строка-пример (серым, помечена «ПРИМЕР — удалите эту строку») и лист
     «Справочники» с допустимыми компаниями/отделами/графиками;
  2. `parse_import_file` — «умный» разбор файла: нормализация значений
     (`5\\2` → `5/2`, `ООО "Комфорт"` → компания справочника, «50 000» → число)
     и валидация каждой строки для превью, БЕЗ записи в БД;
  3. `import_valid_rows` — создание сотрудников по валидным строкам после
     подтверждения (через `build_employee`, как обычный CRUD).

Колонки фиксированные и разбираются ПОЗИЦИОННО (см. `COLUMNS`) — заголовки
нужны человеку, парсер на них не опирается.
"""
from __future__ import annotations

import datetime
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.audit import log_action
from app.models.companies import Company
from app.models.departments import Department
from app.models.employees import Employee
from app.models.schedules import Schedule
from app.schemas.employee import EmployeeCreate
from app.schemas.employee_import import EmployeeImportResult, ImportRowRead
from app.services.company_order import company_order_by
from app.services.employees import build_employee

# Строка-пример помечается этим текстом в первой колонке; парсер такие строки
# пропускает (пользователь может её и удалить — тогда данные идут со 2-й строки).
EXAMPLE_MARKER = "ПРИМЕР — удалите эту строку"


@dataclass(frozen=True)
class _Column:
    key: str
    title: str
    example: str
    hint: str
    width: int


COLUMNS: tuple[_Column, ...] = (
    _Column("tab_number", "Табельный номер", EXAMPLE_MARKER,
            "необязательно; дубль — ошибка строки", 30),
    _Column("full_name", "ФИО *", "Иванов Иван Иванович",
            "обязательно", 30),
    _Column("company", "Компания (основная) *", 'ООО "Комфорт"',
            "обязательно; должна быть в справочнике (код или название)", 26),
    _Column("department", "Отдел", "ИТО",
            "пусто — без отдела; иначе должен быть в справочнике", 20),
    _Column("position", "Должность", "Инженер",
            "любой текст", 24),
    _Column("schedule", "График", "5/2",
            "должен быть в справочнике: 5/2, 2/2 смена 1 …", 18),
    _Column("pay_type", "Тип оплаты", "окладная",
            "окладная / посменная (по умолчанию окладная)", 16),
    _Column("rate", "Оклад", "50 000",
            "для окладной; число", 14),
    _Column("shift_rate", "Ставка за смену", "",
            "для посменной; число", 16),
    _Column("weekend_pay_type", "Оплата выходных", "коэффициент",
            "коэффициент / фикс (по умолчанию коэффициент)", 18),
    _Column("weekend_value", "Коэффициент / ставка выходных", "1,5",
            "1,5 для коэффициента или 740 для фикс. ставки", 24),
    _Column("hire_date", "Дата приёма", "01.03.2026",
            "ДД.ММ.ГГГГ или дата Excel", 16),
)

_HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
_EXAMPLE_FILL = PatternFill("solid", fgColor="FFF2F2F2")
_HINT_FILL = PatternFill("solid", fgColor="FFFFF9E5")


def generate_import_template(db: Session) -> bytes:
    """Шаблон для заполнения: лист «Сотрудники» + лист «Справочники»."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    header_font = Font(name="Arial", size=10, bold=True)
    example_font = Font(name="Arial", size=10, italic=True, color="FF808080")
    hint_font = Font(name="Arial", size=8, italic=True, color="FF9C6500")
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = col.width

        head = ws.cell(row=1, column=idx, value=col.title)
        head.font = header_font
        head.fill = _HEADER_FILL
        head.alignment = wrap

        example = ws.cell(row=2, column=idx, value=col.example or None)
        example.font = example_font
        example.fill = _EXAMPLE_FILL
        example.alignment = Alignment(vertical="center", wrap_text=True)

        hint = ws.cell(row=3, column=idx, value=col.hint)
        hint.font = hint_font
        hint.fill = _HINT_FILL
        hint.alignment = wrap

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"

    _write_reference_sheet(wb, db)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_reference_sheet(wb: Workbook, db: Session) -> None:
    """Лист со списком допустимых значений справочников — чтобы заполняющий
    не гадал, как называется компания или график (иначе строка будет с ошибкой)."""
    ws = wb.create_sheet("Справочники")
    bold = Font(name="Arial", size=10, bold=True)

    companies = (
        db.query(Company).filter(Company.is_active.is_(True))
        .order_by(*company_order_by()).all()
    )
    departments = (
        db.query(Department).filter(Department.is_active.is_(True)).order_by(Department.name).all()
    )
    schedules = (
        db.query(Schedule).filter(Schedule.is_active.is_(True)).order_by(Schedule.name).all()
    )

    blocks = [
        ("Компании (код — название)", [f"{c.code} — {c.name}" for c in companies]),
        ("Отделы", [d.name for d in departments]),
        ("Графики", [s.name for s in schedules]),
    ]

    for col_idx, (title, values) in enumerate(blocks, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 36
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = bold
        cell.fill = _HEADER_FILL
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"


# ── Нормализация значений ─────────────────────────────────────────────────────

# Пробелы, которые Excel/копипаста приносят вместо обычного: неразрывный, узкий и пр.
_SPACE_CHARS = "\u00a0\u2007\u202f\u2009\u2002\u2003\u2004\u2005\u2006\u2008\u200a\ufeff\t"
_QUOTE_CHARS = '"«»“”„‘’\''
_LEGAL_FORMS = ("ооо", "оао", "зао", "пао", "нао", "ао", "ип")


def cell_text(value: object) -> str:
    """Значение ячейки как текст: спецпробелы → обычные, схлопнуть, обрезать."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    for ch in _SPACE_CHARS:
        text = text.replace(ch, " ")
    return re.sub(r"\s+", " ", text).strip()


def match_key(value: object) -> str:
    """Ключ для сопоставления со справочником: регистр и «ё» не различаем."""
    return cell_text(value).lower().replace("ё", "е")


def normalize_schedule_key(value: object) -> str:
    """`5\\2`, `5 / 2`, ` 5/2 ` → `5/2` (регистронезависимо)."""
    key = match_key(value).replace("\\", "/")
    return re.sub(r"\s*/\s*", "/", key)


def company_keys(value: object) -> list[str]:
    """Варианты ключа компании: как есть, без кавычек, без правовой формы.

    `ООО "Комфорт"` даёт ключи `ооо "комфорт"`, `ооо комфорт`, `комфорт` —
    поэтому сопоставляется и с полным названием в справочнике, и с коротким.
    """
    base = match_key(value)
    if not base:
        return []
    keys = [base]

    unquoted = base
    for ch in _QUOTE_CHARS:
        unquoted = unquoted.replace(ch, " ")
    unquoted = re.sub(r"\s+", " ", unquoted).strip()
    if unquoted and unquoted not in keys:
        keys.append(unquoted)

    for form in _LEGAL_FORMS:
        if unquoted.startswith(f"{form} "):
            short = unquoted[len(form) + 1:].strip()
            if short and short not in keys:
                keys.append(short)
            break

    return keys


def parse_decimal(value: object) -> Decimal | None:
    """«50 000», «50 000,50», 50000.0 → Decimal. Пусто → None, мусор → ValueError."""
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError("не число")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    text = cell_text(value)
    if not text:
        return None
    text = text.replace(" ", "")
    text = re.sub(r"(?i)(руб\.?|р\.|₽)$", "", text)
    text = text.replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError("не число") from exc


_DATE_FORMATS = ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y.%m.%d")
# Excel хранит даты числом дней от 30.12.1899 (учитывая его «1900 високосный»).
_EXCEL_EPOCH = datetime.date(1899, 12, 30)


def parse_date(value: object) -> datetime.date | None:
    """Дата из ячейки: объект даты, число Excel или строка в частых форматах."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, bool):
        raise ValueError("не дата")
    if isinstance(value, (int, float)):
        serial = int(value)
        if serial <= 0:
            raise ValueError("не дата")
        return _EXCEL_EPOCH + datetime.timedelta(days=serial)

    text = cell_text(value)
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("не дата")


def parse_pay_type(value: object) -> str | None:
    """«окладная»/«оклад»/«salary» → salary; «посменная»/«за смену» → per_shift.

    Пусто → salary (дефолт), непонятное → None (ошибка строки).
    """
    key = match_key(value)
    if not key:
        return "salary"
    if "смен" in key or "shift" in key:
        return "per_shift"
    if "оклад" in key or "salary" in key:
        return "salary"
    return None


def parse_weekend_pay_type(value: object) -> str | None:
    """«коэффициент»/«коэф»/«×1.5» → coefficient; «фикс»/«740» → fixed_rate.

    Пусто → coefficient (дефолт), непонятное → None (ошибка строки).
    """
    key = match_key(value)
    if not key:
        return "coefficient"
    if "фикс" in key or "fix" in key:
        return "fixed_rate"
    if "коэф" in key or "coef" in key or key.startswith(("×", "x", "*")):
        return "coefficient"
    # Голое число: 1,5 — это коэффициент, 740 — ставка за час.
    try:
        number = parse_decimal(key)
    except ValueError:
        return None
    if number is None:
        return "coefficient"
    return "fixed_rate" if number >= 10 else "coefficient"


# ── Разбор и валидация файла ──────────────────────────────────────────────────

# Верхняя граница на размер файла: 300+ строк — рабочий случай, десятки тысяч —
# почти наверняка не тот файл.
MAX_IMPORT_ROWS = 5000

_SHEET_NAME = "Сотрудники"


class ImportFileError(Exception):
    """Файл целиком непригоден (не .xlsx, нет строк, слишком большой)."""


@dataclass(frozen=True)
class _Refs:
    """Справочники, разложенные по ключам сопоставления (собираются один раз)."""

    companies: dict[str, Company]
    departments: dict[str, Department]
    schedules: dict[str, Schedule]


def _plain_keys(value: object) -> list[str]:
    """Ключи без разбора правовой формы: как есть и без кавычек."""
    base = match_key(value)
    if not base:
        return []
    unquoted = base
    for ch in _QUOTE_CHARS:
        unquoted = unquoted.replace(ch, " ")
    unquoted = re.sub(r"\s+", " ", unquoted).strip()
    return [base] if unquoted == base else [base, unquoted]


def _load_refs(db: Session) -> _Refs:
    companies: dict[str, Company] = {}
    for company in db.query(Company).filter(Company.is_active.is_(True)).all():
        for key in [*company_keys(company.name), *_plain_keys(company.code)]:
            companies.setdefault(key, company)

    departments: dict[str, Department] = {}
    for dept in db.query(Department).filter(Department.is_active.is_(True)).all():
        for key in [*_plain_keys(dept.name), *_plain_keys(dept.code)]:
            departments.setdefault(key, dept)

    schedules: dict[str, Schedule] = {}
    for schedule in db.query(Schedule).filter(Schedule.is_active.is_(True)).all():
        key = normalize_schedule_key(schedule.name)
        if key:
            schedules.setdefault(key, schedule)

    return _Refs(companies=companies, departments=departments, schedules=schedules)


def _is_example_row(raw: dict[str, str]) -> bool:
    """Строка-пример из шаблона — не данные, её не импортируем."""
    if match_key(raw.get("tab_number")).startswith("пример"):
        return True
    marker = match_key(EXAMPLE_MARKER)
    return any(match_key(value) == marker for value in raw.values())


def _parse_row(
    row_number: int,
    raw: dict[str, str],
    refs: _Refs,
    taken_tab_numbers: set[str],
) -> ImportRowRead:
    """Одна строка файла → распознанные значения + список ошибок."""
    errors: list[str] = []

    tab_number = raw["tab_number"] or None
    full_name = raw["full_name"] or None
    position = raw["position"] or None

    if not full_name:
        errors.append("ФИО обязательно")

    if tab_number and match_key(tab_number) in taken_tab_numbers:
        errors.append(f"Таб.№ «{tab_number}» уже существует")

    # Компания — обязательна и должна быть в справочнике
    company = None
    if not raw["company"]:
        errors.append("Компания обязательна")
    else:
        company = _lookup(refs.companies, company_keys(raw["company"]))
        if company is None:
            errors.append(f"Компания «{raw['company']}» не найдена")

    # Отдел — необязателен, но если указан, должен существовать
    department = None
    if raw["department"]:
        department = _lookup(refs.departments, _plain_keys(raw["department"]))
        if department is None:
            errors.append(f"Отдел «{raw['department']}» не найден")

    # График — необязателен, но если указан, должен существовать
    schedule = None
    if raw["schedule"]:
        schedule = _lookup(refs.schedules, [normalize_schedule_key(raw["schedule"])])
        if schedule is None:
            errors.append(f"График «{raw['schedule']}» не найден")

    pay_type = parse_pay_type(raw["pay_type"])
    if pay_type is None:
        errors.append(f"Неизвестный тип оплаты «{raw['pay_type']}»")
        pay_type = "salary"

    rate, rate_ok = _decimal_or_error(raw["rate"], "Оклад", errors)
    shift_rate, shift_rate_ok = _decimal_or_error(raw["shift_rate"], "Ставка за смену", errors)

    # Оклад и ставка за смену взаимоисключающие — чужое поле не переносим.
    # «не число» уже отмечено выше, второй ошибкой про «не указан» не сорим.
    if pay_type == "per_shift":
        rate = None
        if shift_rate is None:
            if shift_rate_ok:
                errors.append("Не указана ставка за смену")
        elif shift_rate <= 0:
            errors.append("Ставка за смену должна быть больше 0")
    else:
        shift_rate = None
        if rate is None:
            if rate_ok:
                errors.append("Не указан оклад")
        elif rate <= 0:
            errors.append("Оклад должен быть больше 0")

    weekend_pay_type = parse_weekend_pay_type(raw["weekend_pay_type"])
    if weekend_pay_type is None:
        errors.append(f"Неизвестный вид оплаты выходных «{raw['weekend_pay_type']}»")
        weekend_pay_type = "coefficient"

    weekend_value, _ = _decimal_or_error(
        raw["weekend_value"], "Коэффициент / ставка выходных", errors
    )
    weekend_coefficient = None
    weekend_fixed_rate = None
    if weekend_pay_type == "fixed_rate":
        weekend_fixed_rate = weekend_value
        if weekend_value is None:
            errors.append("Не указана фиксированная ставка за выходные")
    else:
        # Пусто → дефолт 1.5, как в карточке сотрудника
        weekend_coefficient = weekend_value if weekend_value is not None else Decimal("1.5")

    hire_date = None
    if raw["hire_date"]:
        try:
            hire_date = parse_date(raw["hire_date"])
        except ValueError:
            errors.append(f"Дата приёма «{raw['hire_date']}» не распознана")

    return ImportRowRead(
        row_number=row_number,
        is_valid=not errors,
        errors=errors,
        raw=raw,
        tab_number=tab_number,
        full_name=full_name,
        position=position,
        company_id=company.id if company else None,
        company_name=company.name if company else None,
        department_id=department.id if department else None,
        department_name=department.name if department else None,
        schedule_id=schedule.id if schedule else None,
        schedule_name=schedule.name if schedule else None,
        pay_type=pay_type,
        rate=rate,
        shift_rate=shift_rate,
        weekend_pay_type=weekend_pay_type,
        weekend_coefficient=weekend_coefficient,
        weekend_fixed_rate=weekend_fixed_rate,
        hire_date=hire_date,
    )


def _lookup(index: dict[str, object], keys: list[str]):
    for key in keys:
        found = index.get(key)
        if found is not None:
            return found
    return None


def _decimal_or_error(text: str, label: str, errors: list[str]) -> tuple[Decimal | None, bool]:
    """Число из ячейки. Второй элемент — «разобралось» (пустая ячейка тоже ок)."""
    if not text:
        return None, True
    try:
        return parse_decimal(text), True
    except ValueError:
        errors.append(f"{label} не число: «{text}»")
        return None, False


def parse_import_file(db: Session, content: bytes) -> EmployeeImportResult:
    """Разобрать и провалидировать файл. В БД ничего не пишет — это превью."""
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl бросает разное на битых файлах
        raise ImportFileError(
            "Не удалось прочитать файл. Нужен .xlsx (шаблон можно скачать кнопкой рядом)."
        ) from exc

    try:
        ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.worksheets[0]

        # Таб.№, которые уже заняты: в БД + встреченные выше по файлу
        taken_tab_numbers = {
            match_key(value)
            for (value,) in db.query(Employee.tab_number).filter(
                Employee.tab_number.isnot(None)
            )
            if match_key(value)
        }

        refs = _load_refs(db)
        rows: list[ImportRowRead] = []

        for excel_row, values in enumerate(
            ws.iter_rows(min_row=2, max_col=len(COLUMNS), values_only=True), start=2
        ):
            raw = {col.key: cell_text(value) for col, value in zip(COLUMNS, values)}
            if not any(raw.values()):
                continue
            if _is_example_row(raw):
                continue
            if len(rows) >= MAX_IMPORT_ROWS:
                raise ImportFileError(
                    f"В файле больше {MAX_IMPORT_ROWS} строк — похоже, это не список сотрудников."
                )

            row = _parse_row(excel_row, raw, refs, taken_tab_numbers)
            # Таб.№ занимает только валидная строка: ошибочная не импортируется,
            # значит номер остаётся свободным для следующих.
            if row.is_valid and row.tab_number:
                taken_tab_numbers.add(match_key(row.tab_number))
            rows.append(row)
    finally:
        wb.close()

    if not rows:
        raise ImportFileError("В файле нет строк с данными (заполните шаблон с 3-й строки).")

    valid_count = sum(1 for row in rows if row.is_valid)
    return EmployeeImportResult(
        confirmed=False,
        total=len(rows),
        valid_count=valid_count,
        error_count=len(rows) - valid_count,
        rows=rows,
    )


def import_valid_rows(
    db: Session, actor: Employee, result: EmployeeImportResult
) -> EmployeeImportResult:
    """Создать сотрудников по валидным строкам превью. Ошибочные пропускаются.

    Карточка собирается тем же `build_employee`, что и обычное создание, доступы
    (email/роль/пароль) не импортируются. Всё в одной транзакции: при сбое не
    останется половины сотрудников.
    """
    created = 0
    for row in result.rows:
        if not row.is_valid:
            continue

        payload = EmployeeCreate(
            tab_number=row.tab_number,
            full_name=row.full_name,
            position=row.position,
            department_id=row.department_id,
            schedule_id=row.schedule_id,
            default_company_id=row.company_id,
            pay_type=row.pay_type,
            rate=row.rate,
            shift_rate=row.shift_rate,
            weekend_pay_type=row.weekend_pay_type,
            weekend_coefficient=row.weekend_coefficient,
            weekend_fixed_rate=row.weekend_fixed_rate,
            hire_date=row.hire_date,
        )
        emp = build_employee(payload)
        db.add(emp)
        db.flush()

        log_action(
            db, actor, "employee", emp.id, "create",
            after={
                "source": "excel_import",
                "tab_number": emp.tab_number,
                "full_name": emp.full_name,
                "default_company_id": emp.default_company_id,
                "department_id": emp.department_id,
                "schedule_id": emp.schedule_id,
                "pay_type": emp.pay_type,
                "rate": str(emp.rate) if emp.rate is not None else None,
                "shift_rate": str(emp.shift_rate) if emp.shift_rate is not None else None,
            },
        )
        row.created = True
        row.employee_id = emp.id
        created += 1

    log_action(
        db, actor, "employee", None, "employees_imported",
        after={"created": created, "skipped": result.error_count, "total": result.total},
    )
    db.commit()

    result.confirmed = True
    result.created_count = created
    result.skipped_count = result.error_count
    return result
