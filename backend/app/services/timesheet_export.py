from __future__ import annotations

import calendar as _cal
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.companies import Company
from app.models.employees import Employee
from app.models.production_calendars import ProductionCalendar
from app.services.calendar import get_month_data, parse_days_string
from app.services.timesheet import get_month_entries, visible_employees_for_actor

_ORG_NAME = "ДЕВЕЛОПМЕНТ ГРУППА «ЗЕМЛЯ МО»"

# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin_side() -> Side:
    return Side(style="thin", color="000000")


def _thin_border() -> Border:
    s = _thin_side()
    return Border(left=s, right=s, top=s, bottom=s)


def _holiday_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FFFFCCCC")  # ARGB: red tint


def _short_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FFFFF2CC")  # ARGB: yellow tint


def _weekend_fill() -> PatternFill:
    return PatternFill("solid", fgColor="FFEFEFEF")  # ARGB: gray tint


def _zebra_fill() -> PatternFill:
    """Чередующаяся заливка блока сотрудника — для читаемости при печати."""
    return PatternFill("solid", fgColor="FFF4F7FB")  # очень светлый голубой


def _header_font(bold: bool = True) -> Font:
    return Font(name="Arial", size=9, bold=bold)


def _set_cell(ws, row: int, col: int, value=None, *, bold=False, center=False,
              wrap=False, fill=None, font_color="000000") -> None:
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", size=9, bold=bold, color=font_color)
    alignment_kwargs: dict = {"wrap_text": wrap}
    if center:
        alignment_kwargs["horizontal"] = "center"
        alignment_kwargs["vertical"] = "center"
    c.alignment = Alignment(**alignment_kwargs)
    c.border = _thin_border()
    if fill is not None:
        c.fill = fill


# ── Column layout (задача: структура столбцов по образцу финдира) ──────────────
#
# Порядок фиксированных колонок слева:
#   № → Таб.№ → ФИО → Компания → Подразделение → Должность → График
# Затем дни месяца (1-я / 2-я половина, без промежуточных подытогов), затем
# сводные колонки:
#   Кол-во дней отпуска → Кол-во дней больничного → Норма дней → Факт дней →
#   НОРМА ч/мес → ФАКТ ч/мес → Итого Ч компании → Сверхур. Ч → Праздн. Ч →
#   Итого Ч сотруд.
#
# Компания, дни, Итого Ч компании, Сверхур. Ч, Праздн. Ч — per-row (на каждую
# компанию сотрудника). Остальное — employee-level (merge по строкам компаний).
_COL_NUM = 1
_COL_TAB = 2
_COL_NAME = 3
_COL_COMPANY = 4
_COL_DEPT = 5
_COL_POS = 6
_COL_SCHED = 7
_FIXED_COLS = 7


def _day_col(day: int, total_days: int) -> int:
    """Column index for a calendar day (1-based day → column index).
    Подытогов 1-15 / 16-конец больше нет: дни идут сплошняком."""
    return _FIXED_COLS + day


def _vac_col(total_days: int) -> int:
    """Кол-во дней отпуска (employee-level, задел на будущее — пока пусто)."""
    return _FIXED_COLS + total_days + 1


def _sick_col(total_days: int) -> int:
    """Кол-во дней больничного (employee-level, задел на будущее — пока пусто)."""
    return _FIXED_COLS + total_days + 2


def _norm_days_col(total_days: int) -> int:
    """Норма дней (employee-level)."""
    return _FIXED_COLS + total_days + 3


def _fact_days_col(total_days: int) -> int:
    """Факт дней — кол-во дней с часами (employee-level)."""
    return _FIXED_COLS + total_days + 4


def _norm_col(total_days: int) -> int:
    """НОРМА кол-во раб. часов в месяце (employee-level, merge)."""
    return _FIXED_COLS + total_days + 5


def _fact_hours_col(total_days: int) -> int:
    """ФАКТ кол-во отработанных часов в месяце (employee-level, merge)."""
    return _FIXED_COLS + total_days + 6


def _total_col(total_days: int) -> int:
    """Итого Ч этой компании (per-row, итог компании за месяц)."""
    return _FIXED_COLS + total_days + 7


def _ot_hours_col(total_days: int) -> int:
    """Сверхур. Ч по компании (per-row)."""
    return _FIXED_COLS + total_days + 8


def _hol_hours_col(total_days: int) -> int:
    """Праздн. Ч по компании (per-row)."""
    return _FIXED_COLS + total_days + 9


def _grand_total_col(total_days: int) -> int:
    """Общий Итого Ч сотрудника (merge по строкам компаний)."""
    return _FIXED_COLS + total_days + 10


def _last_col(total_days: int) -> int:
    """Самая правая колонка таблицы."""
    return _grand_total_col(total_days)


def _distribute_int(total: int, weights: dict[int, float]) -> dict[int, int]:
    """Целочисленное распределение total по ключам пропорционально weights
    (метод наибольших остатков). Сумма частей = total."""
    wsum = sum(weights.values())
    if wsum <= 0 or total == 0:
        return {k: 0 for k in weights}
    parts: dict[int, int] = {}
    rem: dict[int, float] = {}
    for k, w in weights.items():
        exact = total * w / wsum
        parts[k] = int(exact)
        rem[k] = exact - int(exact)
    leftover = total - sum(parts.values())
    for k in sorted(weights, key=lambda x: (-rem[x], x))[:leftover]:
        parts[k] += 1
    return parts


# ── Main generator ────────────────────────────────────────────────────────────

def generate_t13_excel(
    db: Session,
    actor: Employee,
    year: int,
    month: int,
    department_id: int | None = None,
) -> bytes:
    """
    Генерирует Excel-файл с табелем Т-13 за указанный период.
    Возвращает bytes — содержимое .xlsx файла.
    """
    employees = visible_employees_for_actor(db, actor, department_id, year=year, month=month)
    entries = get_month_entries(db, employees, year, month)

    # Производственный календарь (опционально — не падаем если нет)
    cal_record = db.query(ProductionCalendar).filter_by(year=year).first()
    calendar_data: dict | None = cal_record.data if cal_record else None

    # Компании
    companies_by_id: dict[int, Company] = {
        c.id: c for c in db.query(Company).filter(Company.is_active == True).all()  # noqa: E712
    }

    # Индекс entries: {(employee_id, day): {company_id: hours}}
    entries_index: dict[tuple[int, int], dict[int, float]] = {}
    for e in entries:
        day = e.work_date.day
        key = (e.employee_id, day)
        if key not in entries_index:
            entries_index[key] = {}
        entries_index[key][e.company_id] = float(e.hours)

    # Для каждого сотрудника — список компаний, по которым есть entries в этом месяце
    emp_companies: dict[int, list[int]] = {}
    for emp in employees:
        used: set[int] = set()
        for e in entries:
            if e.employee_id == emp.id:
                used.add(e.company_id)
        if not used:
            continue  # сотрудник без entries — пропускаем
        # Сортируем по id для детерминизма
        emp_companies[emp.id] = sorted(used)

    total_days = _cal.monthrange(year, month)[1]

    # Определяем типы дней
    non_working: set[int] = set()
    short_days: set[int] = set()
    if calendar_data:
        month_data = get_month_data(calendar_data, month)
        if month_data:
            non_working, short_days = parse_days_string(month_data["days"])
    # Выходные по ISO-weekday (6=Сб, 7=Вс → weekday() 5,6)
    weekends: set[int] = {
        d for d in range(1, total_days + 1)
        if date(year, month, d).weekday() >= 5
    }

    wb = Workbook()
    ws = wb.active
    ws.title = f"Т-13 {year}-{month:02d}"

    # ── Ширины колонок ───────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(_COL_NUM)].width = 5
    ws.column_dimensions[get_column_letter(_COL_TAB)].width = 11
    ws.column_dimensions[get_column_letter(_COL_NAME)].width = 28
    ws.column_dimensions[get_column_letter(_COL_COMPANY)].width = 22
    ws.column_dimensions[get_column_letter(_COL_DEPT)].width = 14
    ws.column_dimensions[get_column_letter(_COL_POS)].width = 18
    ws.column_dimensions[get_column_letter(_COL_SCHED)].width = 7
    for d in range(1, total_days + 1):
        ws.column_dimensions[get_column_letter(_day_col(d, total_days))].width = 4
    for col_fn, width in (
        (_vac_col, 8), (_sick_col, 8), (_norm_days_col, 7), (_fact_days_col, 7),
        (_norm_col, 9), (_fact_hours_col, 9), (_total_col, 8),
        (_ot_hours_col, 8), (_hol_hours_col, 8), (_grand_total_col, 9),
    ):
        ws.column_dimensions[get_column_letter(col_fn(total_days))].width = width

    # ── Шапка документа ─────────────────────────────────────────────────────
    cur_row = _write_document_header(ws, year, month, department_id, db, total_days)

    # ── Шапка таблицы ───────────────────────────────────────────────────────
    cur_row = _write_table_header(ws, cur_row, year, month, total_days,
                                  non_working, short_days, weekends)

    # ── Строки сотрудников ──────────────────────────────────────────────────
    seq = 0
    for emp in employees:
        if emp.id not in emp_companies:
            continue
        company_ids = emp_companies[emp.id]
        seq += 1
        cur_row = _write_employee_rows(
            ws, cur_row, seq, emp, company_ids, companies_by_id,
            entries_index, total_days,
            non_working, short_days, weekends,
        )

    # ── Подвал с подписями ──────────────────────────────────────────────────
    _write_footer(ws, cur_row + 2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Document header ───────────────────────────────────────────────────────────

def _write_document_header(
    ws, year: int, month: int, department_id: int | None, db: Session, total_days: int,
) -> int:
    """Пишет шапку документа, возвращает следующую строку."""

    # Определяем подразделение
    dept_name = "Все отделы"
    if department_id is not None:
        from app.models.departments import Department
        dept = db.get(Department, department_id)
        if dept:
            dept_name = dept.name

    last_day = _cal.monthrange(year, month)[1]
    period_start = f"01.{month:02d}.{year}"
    period_end = f"{last_day:02d}.{month:02d}.{year}"

    right_col = _last_col(total_days)

    # Строка 1: «Унифицированная форма...»
    ws.merge_cells(start_row=1, start_column=right_col - 3,
                   end_row=1, end_column=right_col)
    ws.cell(row=1, column=right_col - 3).value = "Унифицированная форма № Т-13"
    ws.cell(row=1, column=right_col - 3).font = Font(name="Arial", size=8)

    # Строка 2: название организации
    ws.cell(row=2, column=1).value = _ORG_NAME
    ws.cell(row=2, column=1).font = Font(name="Arial", size=10, bold=True)

    # Строка 3: «Форма по ОКУД»
    ws.merge_cells(start_row=3, start_column=right_col - 3,
                   end_row=3, end_column=right_col)
    ws.cell(row=3, column=right_col - 3).value = "Форма по ОКУД: 0301008"
    ws.cell(row=3, column=right_col - 3).font = Font(name="Arial", size=8)

    # Строка 4: подразделение
    ws.cell(row=4, column=1).value = f"Структурное подразделение: {dept_name}"
    ws.cell(row=4, column=1).font = Font(name="Arial", size=9)

    # Строка 5: ТАБЕЛЬ
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=right_col)
    c = ws.cell(row=6, column=1)
    c.value = "ТАБЕЛЬ УЧЁТА РАБОЧЕГО ВРЕМЕНИ"
    c.font = Font(name="Arial", size=14, bold=True)
    c.alignment = Alignment(horizontal="center")

    # Строка 7: период
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=right_col)
    c = ws.cell(row=7, column=1)
    c.value = f"Отчётный период: с {period_start} по {period_end}"
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal="center")

    ws.row_dimensions[8].height = 6  # пустая строка-разделитель

    return 9  # первая строка таблицы


# ── Table header ──────────────────────────────────────────────────────────────

def _write_table_header(
    ws, start_row: int, year: int, month: int, total_days: int,
    non_working: set[int], short_days: set[int], weekends: set[int],
) -> int:
    r = start_row
    weekday_ru = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    def _merged_header(col: int, label: str) -> None:
        ws.merge_cells(start_row=r, start_column=col, end_row=r + 1, end_column=col)
        c = ws.cell(row=r, column=col)
        c.value = label
        c.font = _header_font()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
        ws.cell(row=r + 1, column=col).border = _thin_border()

    # Заголовки фиксированных колонок — объединяем 2 строки
    for col, label in (
        (_COL_NUM, "№"),
        (_COL_TAB, "Таб. №"),
        (_COL_NAME, "ФИО"),
        (_COL_COMPANY, "Компания"),
        (_COL_DEPT, "Подразделение"),
        (_COL_POS, "Должность"),
        (_COL_SCHED, "График"),
    ):
        _merged_header(col, label)

    # День 1-15: строка 1 — «1-я половина месяца», строка 2 — числа
    first_half_start = _day_col(1, total_days)
    first_half_end = _day_col(15, total_days)
    ws.merge_cells(start_row=r, start_column=first_half_start,
                   end_row=r, end_column=first_half_end)
    c = ws.cell(row=r, column=first_half_start)
    c.value = "1-я половина"
    c.font = _header_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin_border()

    # День 16-end: строка 1 — «2-я половина»
    second_half_start = _day_col(16, total_days)
    second_half_end = _day_col(total_days, total_days)
    ws.merge_cells(start_row=r, start_column=second_half_start,
                   end_row=r, end_column=second_half_end)
    c = ws.cell(row=r, column=second_half_start)
    c.value = "2-я половина"
    c.font = _header_font()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin_border()

    # Сводные employee-/per-company колонки (2-строчные заголовки)
    for col, label in (
        (_vac_col(total_days), "Кол-во дней\nотпуска"),
        (_sick_col(total_days), "Кол-во дней\nбольничного"),
        (_norm_days_col(total_days), "Норма\nдней"),
        (_fact_days_col(total_days), "Факт\nдней"),
        (_norm_col(total_days), "НОРМА\nч/мес"),
        (_fact_hours_col(total_days), "ФАКТ\nч/мес"),
        (_total_col(total_days), "Итого Ч\nкомпании"),
        (_ot_hours_col(total_days), "Сверхур.\nЧ"),
        (_hol_hours_col(total_days), "Праздн.\nЧ"),
        (_grand_total_col(total_days), "Итого Ч\nсотруд."),
    ):
        _merged_header(col, label)

    # Вторая строка заголовка: числа дней
    for d in range(1, total_days + 1):
        col = _day_col(d, total_days)
        fill = None
        font_color = "000000"
        if d in non_working:
            fill = _holiday_fill()
            font_color = "CC0000"
        elif d in short_days:
            fill = _short_fill()
            font_color = "7D6608"
        elif d in weekends:
            fill = _weekend_fill()
            font_color = "666666"

        wd = date(year, month, d).weekday()
        label = f"{d}\n{weekday_ru[wd]}"
        c = ws.cell(row=r + 1, column=col)
        c.value = label
        c.font = Font(name="Arial", size=8, bold=True, color=font_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
        if fill:
            c.fill = fill

    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r + 1].height = 26
    return r + 2


# ── Employee rows ─────────────────────────────────────────────────────────────

def _write_employee_rows(
    ws,
    start_row: int,
    seq: int,
    emp: Employee,
    company_ids: list[int],
    companies_by_id: dict[int, Company],
    entries_index: dict[tuple[int, int], dict[int, float]],
    total_days: int,
    non_working: set[int] | None = None,
    short_days: set[int] | None = None,
    weekends: set[int] | None = None,
) -> int:
    non_working = non_working or set()
    short_days = short_days or set()
    weekends = weekends or set()
    off_days = non_working | weekends  # праздники/выходные → праздничные часы

    n = len(company_ids)
    end_row = start_row + n - 1

    # Помесячные итоги/праздничные часы по компаниям этого сотрудника
    comp_totals: dict[int, float] = {}
    comp_holiday: dict[int, float] = {}

    def _fmt(v: float):
        return int(v) if v == int(v) else round(v, 2)

    # ── Employee-level колонки (merge по всем строкам сотрудника) ─────────────
    schedule = emp.schedule
    is_standard = schedule is not None and schedule.schedule_type != "shift"
    norm_days = (
        sum(1 for d in range(1, total_days + 1) if d not in off_days)
        if is_standard else None
    )
    fact_days = sum(
        1 for d in range(1, total_days + 1)
        if any(h > 0 for h in entries_index.get((emp.id, d), {}).values())
    )
    norm_hours = _employee_norm_hours(emp, total_days, short_days, off_days)

    def _emp_cell(col: int, value, *, bold=False, left=False):
        if n > 1:
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)
        c = ws.cell(row=start_row, column=col)
        if value is not None and value != "":
            c.value = value
        c.font = _header_font(bold=bold)
        horizontal = "left" if left else "center"
        c.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=left)
        c.border = _thin_border()
        for rr in range(start_row + 1, end_row + 1):
            ws.cell(row=rr, column=col).border = _thin_border()

    _emp_cell(_COL_NUM, seq)
    _emp_cell(_COL_TAB, emp.tab_number or "")
    _emp_cell(_COL_NAME, emp.full_name, left=True)
    _emp_cell(_COL_DEPT, emp.department.name if emp.department else "", left=True)
    _emp_cell(_COL_POS, emp.position or "", left=True)
    _emp_cell(_COL_SCHED, emp.schedule.name if emp.schedule else "")
    # Отпуск / больничный — задел на будущее, данных пока нет → пусто.
    _emp_cell(_vac_col(total_days), None)
    _emp_cell(_sick_col(total_days), None)
    _emp_cell(_norm_days_col(total_days), norm_days if norm_days else None)
    _emp_cell(_fact_days_col(total_days), fact_days if fact_days else None)
    _emp_cell(_norm_col(total_days), _fmt(norm_hours) if norm_hours else None)

    # Строки по компаниям (per-row): Компания, дни, Итого Ч компании
    for i, comp_id in enumerate(company_ids):
        row = start_row + i
        comp = companies_by_id.get(comp_id)
        comp_name = comp.name if comp else f"Компания #{comp_id}"

        # Название компании
        c = ws.cell(row=row, column=_COL_COMPANY)
        c.value = comp_name
        c.font = _header_font(bold=False)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _thin_border()

        total_hours = 0.0
        holiday_hours = 0.0

        for d in range(1, total_days + 1):
            col = _day_col(d, total_days)
            hours = entries_index.get((emp.id, d), {}).get(comp_id)
            c = ws.cell(row=row, column=col)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=9)
            if hours is not None and hours > 0:
                c.value = int(hours) if hours == int(hours) else hours
                total_hours += hours
                if d in off_days:
                    holiday_hours += hours

        # Итого Ч компании (per-row)
        c = ws.cell(row=row, column=_total_col(total_days))
        c.value = int(total_hours) if total_hours == int(total_hours) else total_hours
        c.font = _header_font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()

        comp_totals[comp_id] = total_hours
        comp_holiday[comp_id] = holiday_hours

    # ── Переработка (employee-level): ПОМЕСЯЧНО (задача 3.11b п.0) ─────────────
    # переработка = факт_будних_часов − месячная_норма (если положительно).
    # Праздничные/выходные часы — отдельная категория, в переработку не входят.
    overtime = 0
    if is_standard:
        regular_hours = sum(comp_totals.values()) - sum(comp_holiday.values())
        overtime = max(0, int(regular_hours - norm_hours))

    # Переработка по компаниям — пропорционально часам (метод наибольших остатков)
    ot_weights = {cid: comp_totals.get(cid, 0.0) for cid in company_ids}
    comp_overtime = _distribute_int(overtime, ot_weights)

    # ── Per-row: Сверхур. Ч и Праздн. Ч по компании ──────────────────────────
    for i, comp_id in enumerate(company_ids):
        row = start_row + i
        ot = comp_overtime.get(comp_id, 0)
        c = ws.cell(row=row, column=_ot_hours_col(total_days))
        c.value = ot if ot > 0 else None
        c.font = _header_font(bold=False)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()

        hol = comp_holiday.get(comp_id, 0.0)
        c = ws.cell(row=row, column=_hol_hours_col(total_days))
        c.value = _fmt(hol) if hol > 0 else None
        c.font = _header_font(bold=False)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()

    # ── Employee-level (merge): ФАКТ ч/мес и Итого Ч сотрудника ───────────────
    grand_total = sum(comp_totals.values())
    _emp_cell(_fact_hours_col(total_days),
              _fmt(grand_total) if grand_total > 0 else None)
    _emp_cell(_grand_total_col(total_days),
              _fmt(grand_total) if grand_total > 0 else None, bold=True)

    # ── Подсветка блока сотрудника (задача 3.11a п.6) ─────────────────────────
    # Чередующаяся заливка инфо/итоговых колонок для читаемости. Дни не трогаем,
    # чтобы не перекрывать подсветку праздников/сокращённых.
    if seq % 2 == 0:
        fill = _zebra_fill()
        info_cols = [
            _COL_NUM, _COL_TAB, _COL_NAME, _COL_COMPANY, _COL_DEPT,
            _COL_POS, _COL_SCHED,
            _vac_col(total_days), _sick_col(total_days),
            _norm_days_col(total_days), _fact_days_col(total_days),
            _norm_col(total_days), _fact_hours_col(total_days),
            _total_col(total_days), _ot_hours_col(total_days),
            _hol_hours_col(total_days), _grand_total_col(total_days),
        ]
        for rr in range(start_row, end_row + 1):
            for col in info_cols:
                c = ws.cell(row=rr, column=col)
                if c.fill is None or c.fill.fgColor.rgb in (None, "00000000"):
                    c.fill = fill

    return end_row + 1


def _employee_norm_hours(
    emp: Employee, total_days: int, short_days: set[int], off_days: set[int]
) -> float:
    """Месячная норма часов: рабочих дней × hours_per_shift − сокращённые дни.
    Только для standard-графиков; иначе 0."""
    schedule = emp.schedule
    if schedule is None or schedule.schedule_type == "shift":
        return 0.0
    norm = schedule.hours_per_shift
    total = 0.0
    for d in range(1, total_days + 1):
        if d in off_days:
            continue
        total += (norm - 1) if d in short_days else norm
    return total


# ── Footer ────────────────────────────────────────────────────────────────────

def _write_footer(ws, start_row: int) -> None:
    lines = [
        ("Ответственное лицо:", "должность", "подпись", "расшифровка подписи"),
        ("Руководитель:", "должность", "подпись", "расшифровка подписи"),
        ("Работник кадровой службы:", "должность", "подпись", "расшифровка подписи"),
    ]
    r = start_row
    for title, *labels in lines:
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9)
        # Три линии для подписей
        for k, lbl in enumerate(labels):
            col = 3 + k * 3
            ws.cell(row=r, column=col).value = "______________________________"
            ws.cell(row=r, column=col).font = Font(name="Arial", size=9)
            ws.cell(row=r + 1, column=col).value = lbl
            ws.cell(row=r + 1, column=col).font = Font(name="Arial", size=8, color="666666")
            ws.cell(row=r + 1, column=col).alignment = Alignment(horizontal="center")
        r += 3  # строка с линиями + подписи + пустая между блоками
