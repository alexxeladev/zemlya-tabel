"""
Excel-выгрузка сводной ведомости «Расчёт ЗП».

Формат — по образцу финдиректора (task_vedomost_format ч.3, лист «Секьюрити»):
шапка «организация / подразделение / ВЕДОМОСТЬ за месяц», строка ИТОГО СВЕРХУ
(так в образце — бухгалтер видит суммы, не прокручивая лист), строки рабочих
мест, строка ИТОГО снизу.

Состав колонок 1..N+2 повторяет образец и покрывает всё, что видно в веб-
ведомости. То, чего в образце нет (обоснования по видам начислений и ночные
смены), идёт ХВОСТОВЫМИ колонками после «ПРИМЕЧАНИЕ»: в шаблоне финдира буквы
колонок зафиксированы, и вставка в середину сдвинула бы его формулы.

Порядок колонок юрлиц — настроенный в справочнике (ч.1), заголовки —
названиями, не кодами (ч.2); и то, и другое приходит готовым в
`statement.companies`.

Суммы здесь НЕ считаются: всё приходит посчитанным в `PayrollStatementRead`.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.payroll_statement import PayrollStatementRead, StatementRow

# Откуда взято распределение по юрлицам (каскад task_distribution_v2 ч.3)
_SOURCE_LABELS = {
    "month": "проценты переопределены на месяц",
    "employee": "проценты из карточки сотрудника",
    "department": "дефолт отдела",
    "hours": "распределено по часам (авто)",
    # Отдел с флагом «распределение по заявкам» (task_hr_applications)
    "applications": "по заявкам на подбор",
}

_MONTHS = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

# Колонка A образца — узкий отступ, данные начинаются с B.
_FIRST_COL = 2

# Колонки 1..27 образца, в его порядке. Дальше — юрлица, «ИТОГО Разбивка»,
# «ПРИМЕЧАНИЕ» и хвост.
_FIXED_HEADERS = [
    "№ п/п",
    "табельный номер",
    "ФИО",
    "Компания",
    "Подразделение",
    "Должность",
    "ОКЛАД",
    "график работы",
    "Кол-во дней отпуска",
    "Кол-во дней больничного",
    "Норма дней",
    "Факт дней",
    "НОРМА Кол-во раб. час в месяце",
    "ФАКТ Кол-во отраб. час в месяце",
    "Учет переработок 0; 1; 1,5",
    "Кол-во переработки, час",
    "Сумма ПЕРЕРАБОТКи",
    "Начислено, оклад",
    "Выплачено отпуск/больничный",
    "Премия Базовая",
    "KPI по выполнению плана",
    "Премия",
    "Основание",
    "Итого начислено",
    "выплачено аванс\\удержано",
    "Сумма к выплате",
    "Разбивка по %",
]

# Хвост после «ПРИМЕЧАНИЕ» — то, чего в образце нет.
_TAIL_HEADERS = [
    "Обоснование премии",
    "Обоснование KPI",
    "Обоснование удержаний",
    "Ночных смен",
    "Надбавка за ночные",
]

# Колонки, которые суммируются в строках ИТОГО, и откуда берётся сумма.
# Складываем только складываемое: дни и часы — да, оклад и коэффициент
# переработки — нет (сумма окладов отдела смысла не имеет).
#
# Денежные итоги берутся из готовых `total_*` схемы — ровно тех, что стоят в
# подвале веб-ведомости: пересчитав их здесь по строкам, мы завели бы вторую
# формулу, которая однажды разойдётся с экраном. Дни и часы своих `total_*`
# не имеют, они складываются по строкам.
_ROW_SUMMED = {8: "vacation_days", 9: "sick_days", 11: "fact_days", 13: "fact_hours",
               15: "overtime_hours", 21: "premium_extra_amount"}
_STATEMENT_TOTALS = {
    16: "total_overtime_amount", 17: "total_base_salary",
    19: "total_premium", 20: "total_kpi", 23: "total_accrued",
    24: "total_deductions", 25: "total_net_payout",
}
# Целочисленные колонки итогов (дни) — без копеек.
_INT_TOTAL_COLS = (8, 9, 11)

_MONEY_FMT = "#,##0.00"
_INT_FMT = "#,##0"
_HEADER_FILL = "FFCCFFCC"   # светло-зелёный, как в образце
_WARN_FILL = "FFFFE5E5"

# Ширины колонок 1..27 (номера — 0-based в _FIXED_HEADERS); остальным — дефолт.
_FIXED_WIDTHS = {
    0: 6, 1: 15, 2: 28, 3: 20, 4: 18, 5: 20, 6: 13, 7: 11,
    8: 10, 9: 11, 10: 9, 11: 9, 12: 13, 13: 14, 14: 11, 15: 11,
    16: 13, 17: 14, 18: 14, 19: 12, 20: 13, 21: 11, 22: 22,
    23: 14, 24: 14, 25: 14, 26: 22,
}
_DEFAULT_WIDTH = 13
_NOTE_WIDTH = 30
_REASON_WIDTH = 30


def _border() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _money(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(Decimal(value))


def _absence_paid(row: StatementRow) -> Decimal:
    """«Выплачено отпуск/больничный» — одна колонка образца на оба вида."""
    return row.vacation_amount + row.sick_amount


def _percent_text(row: StatementRow, names_by_company: dict[int, str]) -> str:
    """«Разбивка по %» — как в образце: «Секьюрити 100%».

    Компании с нулевой долей не пишем: строка и так узкая, а ноль ничего не
    добавляет. Проценты берём из уже посчитанного распределения — второй раз
    их здесь не считаем.
    """
    parts = []
    for d in row.distribution:
        percent = Decimal(d.percent)
        if percent == 0:
            continue
        # 100.00 → «100», 37.21 → «37.21»: хвост нулей в узкой колонке лишний.
        text = f"{percent.normalize():f}"
        parts.append(f"{names_by_company.get(d.company_id, '')} {text}%".strip())
    return ", ".join(parts)


def _pay_base(row: StatementRow) -> Decimal | None:
    """Что показать в колонке «ОКЛАД» для каждого типа оплаты."""
    if row.pay_type == "per_shift":
        return row.shift_rate
    if row.pay_type == "hourly":
        return row.hour_rate
    return row.rate


def _note_text(row: StatementRow) -> str:
    """«ПРИМЕЧАНИЕ» — тип оплаты, уровень каскада распределения и пояснения."""
    note = row.note or ""
    if row.pay_type == "per_shift":
        note = (
            (note + "; " if note else "")
            + f"посменно: {row.base_shifts} смен × ставку"
            + (
                f" (+{row.worked_shifts - row.base_shifts} смен "
                "в выходные/праздники по коэффициенту)"
                if row.worked_shifts > row.base_shifts else ""
            )
        )
    elif row.pay_type == "hourly":
        # Иначе 450 в колонке «ОКЛАД» читается как месячный оклад.
        note = (note + "; " if note else "") + "почасово: в «ОКЛАД» ставка за час"
    source_label = _SOURCE_LABELS.get(row.distribution_source)
    if source_label and row.distribution:
        note = (note + "; " if note else "") + source_label
    # Отдел «по заявкам», но заявок за месяц нет — распределение ушло на каскад,
    # и в выгрузке это должно быть видно (task_hr_applications).
    if row.distribution_note:
        note = (note + "; " if note else "") + row.distribution_note
    return note


def _deduction_reasons(row: StatementRow) -> list[str]:
    lines = list(row.advance_reasons)
    if row.loan_note:
        lines.append(row.loan_note)
    return lines


def _row_values(row: StatementRow) -> list:
    """Значения колонок 1..27 образца для одной строки (рабочего места).

    № п/п проставляется отдельно — он считается по ЛЮДЯМ, а строк у
    совместителя несколько.
    """
    return [
        None,  # № п/п — заполняется вызывающим
        row.tab_number or "",
        row.employee_name,
        row.main_company_name or "",
        row.department_name or "",
        row.position or "",
        # У посменного и почасовика оклада нет — в колонку идёт их ставка
        # (условный оклад посменного служебный, в ведомость его не выносим).
        # Что это за число, сказано в «ПРИМЕЧАНИЕ» — там указан тип оплаты.
        _money(_pay_base(row)),
        row.schedule_name or "",
        row.vacation_days,
        row.sick_days,
        row.norm_days,
        row.fact_days,
        _money(row.norm_hours),
        _money(row.fact_hours),
        _money(row.overtime_coefficient),
        _money(row.overtime_hours),
        _money(row.overtime_amount),
        _money(row.base_salary),
        _money(_absence_paid(row)),
        _money(row.premium_amount),
        _money(row.kpi_amount),
        _money(row.premium_extra_amount),
        "\n".join(list(row.premium_reasons) + list(row.kpi_reasons)),
        _money(row.accrued_total),
        _money(row.deductions),
        _money(row.net_payout),
        None,  # «Разбивка по %» — заполняется вызывающим (нужны имена юрлиц)
    ]


def generate_statement_excel(statement: PayrollStatementRead) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    border = _border()
    bold = Font(name="Arial", size=9, bold=True)
    normal = Font(name="Arial", size=9, bold=False)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    warn_fill = PatternFill("solid", fgColor=_WARN_FILL)

    companies = statement.companies
    # Заголовок колонки — короткое название (ч.2); полное уходит в примечание
    # ячейки, чтобы «ООО "..."» не терялось совсем.
    names_by_company = {c.id: (c.display_name or c.name or c.code) for c in companies}

    n_fixed = len(_FIXED_HEADERS)
    first_company_col = _FIRST_COL + n_fixed
    dist_total_col = first_company_col + len(companies)
    note_col = dist_total_col + 1
    tail_start = note_col + 1
    last_col = tail_start + len(_TAIL_HEADERS) - 1
    (premium_reason_col, kpi_reason_col, deduction_reason_col,
     night_shifts_col, night_amount_col) = range(tail_start, tail_start + 5)

    # ── Шапка образца: организация / подразделение / период ────────────────────
    heading = [
        (statement.organization, True),
        ("(наименование организации)", False),
        (f'"{statement.subdivision}"', False),
        ("(наименование структурного подразделения)", False),
        (
            "ВЕДОМОСТЬ на выплату заработной платы за "
            f"{_MONTHS[statement.month - 1]} {statement.year}",
            False,
        ),
    ]
    for i, (text, is_bold) in enumerate(heading, start=1):
        c = ws.cell(row=i, column=_FIRST_COL, value=text)
        c.font = Font(name="Times New Roman", size=12, bold=is_bold)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Merge по всей ширине таблицы — она зависит от числа юрлиц.
        ws.merge_cells(
            start_row=i, start_column=_FIRST_COL, end_row=i, end_column=last_col
        )

    header_row = 7
    headers = (
        _FIXED_HEADERS
        + [names_by_company[c.id] for c in companies]
        + ["ИТОГО Разбивка", "ПРИМЕЧАНИЕ"]
        + _TAIL_HEADERS
    )
    for i, title in enumerate(headers):
        c = ws.cell(row=header_row, column=_FIRST_COL + i, value=title)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = header_fill
    # Полное название юрлица — примечанием к заголовку его колонки.
    for j, comp in enumerate(companies):
        if comp.name and comp.name != names_by_company[comp.id]:
            ws.cell(row=header_row, column=first_company_col + j).comment = Comment(
                comp.name, "Расчёт ЗП", height=60, width=240,
            )
    ws.row_dimensions[header_row].height = 76

    # ── Итоги СВЕРХУ (так в образце) и снизу — считаются по одним и тем же данным
    totals_top_row = header_row + 1
    first_data_row = totals_top_row + 1

    row_no_by_employee: dict[int, int] = {}
    row = first_data_row
    for r in statement.rows:
        # Нумерация — по ЛЮДЯМ: у совместителя несколько строк с одним номером.
        seq = row_no_by_employee.setdefault(r.employee_id, len(row_no_by_employee) + 1)

        values = _row_values(r)
        values[0] = seq
        values[26] = _percent_text(r, names_by_company)
        for i, val in enumerate(values):
            c = ws.cell(row=row, column=_FIRST_COL + i, value=val)
            c.font = normal
            c.border = border
            # Текстовые колонки — по левому краю, числовые по центру (как в образце)
            c.alignment = left if i in (2, 3, 4, 5, 22, 26) else center
            if isinstance(val, float):
                c.number_format = _MONEY_FMT
            elif isinstance(val, int):
                c.number_format = _INT_FMT

        # Суммы распределения по юрлицам
        amt_by_company = {d.company_id: d.amount for d in r.distribution}
        for j, comp in enumerate(companies):
            # Ноль, а не пустая ячейка: так в образце, и по строке сразу
            # видно, что юрлицо в распределение не вошло.
            c = ws.cell(
                row=row, column=first_company_col + j,
                value=_money(amt_by_company.get(comp.id, Decimal("0"))),
            )
            c.font = normal
            c.border = border
            c.alignment = center
            c.number_format = _MONEY_FMT
        c = ws.cell(row=row, column=dist_total_col, value=_money(r.distribution_total))
        c.font = bold
        c.border = border
        c.alignment = center
        c.number_format = _MONEY_FMT
        # Ручная сумма процентов ≠ 100 — подсветить (авто-доли не трогаем)
        if r.distribution and not r.is_auto_distributed and r.percent_sum != Decimal("100"):
            c.fill = warn_fill

        c = ws.cell(row=row, column=note_col, value=_note_text(r))
        c.font = normal
        c.border = border
        c.alignment = left

        # ── Хвост: обоснования и ночные ───────────────────────────────────────
        for col, lines in (
            (premium_reason_col, r.premium_reasons),
            (kpi_reason_col, r.kpi_reasons),
            (deduction_reason_col, _deduction_reasons(r)),
        ):
            c = ws.cell(row=row, column=col, value="\n".join(lines) if lines else "")
            c.font = normal
            c.border = border
            c.alignment = left
        for col, val in (
            (night_shifts_col, r.night_shifts or ""),
            (night_amount_col, _money(r.night_amount)),
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.font = normal
            c.border = border
            c.alignment = center
            if isinstance(val, float):
                c.number_format = _MONEY_FMT
        # Обоснование удержаний — ещё и примечанием рядом с самой суммой,
        # чтобы не уезжать за ним в хвост листа.
        for col_idx, lines in (
            (24, _deduction_reasons(r)),
        ):
            if lines:
                ws.cell(row=row, column=_FIRST_COL + col_idx).comment = Comment(
                    "\n".join(lines), "Расчёт ЗП",
                    height=len(lines) * 24 + 24, width=280,
                )
        row += 1

    last_data_row = row - 1
    totals_bottom_row = row

    # ── Строки ИТОГО (сверху и снизу — одинаковые) ────────────────────────────
    fixed_totals: dict[int, Decimal] = {
        idx: sum((Decimal(getattr(r, attr)) for r in statement.rows), Decimal("0"))
        for idx, attr in _ROW_SUMMED.items()
    }
    fixed_totals.update(
        {idx: Decimal(getattr(statement, attr)) for idx, attr in _STATEMENT_TOTALS.items()}
    )
    # «Выплачено отпуск/больничный» — одна колонка на два вида начислений.
    fixed_totals[18] = statement.total_vacation_amount + statement.total_sick_amount
    company_totals = {
        c.id: statement.distribution_totals.get(c.id, Decimal("0")) for c in companies
    }
    grand_dist = sum(company_totals.values(), Decimal("0"))
    night_shifts_total = sum(r.night_shifts for r in statement.rows)

    for totals_row, label_col in (
        (totals_top_row, None), (totals_bottom_row, 1),
    ):
        for i in range(len(headers)):
            c = ws.cell(row=totals_row, column=_FIRST_COL + i)
            c.border = border
            c.fill = header_fill
            c.font = bold
            c.alignment = center
        if label_col is not None:
            lc = ws.cell(row=totals_row, column=_FIRST_COL + label_col, value="ИТОГО")
            lc.font = bold
            lc.alignment = center
        for idx, value in fixed_totals.items():
            c = ws.cell(row=totals_row, column=_FIRST_COL + idx, value=_money(value))
            c.number_format = _INT_FMT if idx in _INT_TOTAL_COLS else _MONEY_FMT
        for j, comp in enumerate(companies):
            c = ws.cell(
                row=totals_row, column=first_company_col + j,
                value=_money(company_totals[comp.id]),
            )
            c.number_format = _MONEY_FMT
        ws.cell(
            row=totals_row, column=dist_total_col, value=_money(grand_dist)
        ).number_format = _MONEY_FMT
        ws.cell(
            row=totals_row, column=night_shifts_col, value=night_shifts_total or None
        ).number_format = _INT_FMT
        ws.cell(
            row=totals_row, column=night_amount_col,
            value=_money(statement.total_night_amount),
        ).number_format = _MONEY_FMT

    # ── Ширины колонок и заморозка шапки ──────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 3  # узкий отступ образца
    for i in range(len(headers)):
        col = _FIRST_COL + i
        if i < n_fixed:
            width = _FIXED_WIDTHS.get(i, _DEFAULT_WIDTH)
        elif col == note_col:
            width = _NOTE_WIDTH
        elif col in (premium_reason_col, kpi_reason_col, deduction_reason_col):
            width = _REASON_WIDTH
        else:
            width = _DEFAULT_WIDTH
        ws.column_dimensions[get_column_letter(col)].width = width
    # Заголовки и верхний ИТОГО остаются на экране при прокрутке; ФИО — слева.
    ws.freeze_panes = ws.cell(row=first_data_row, column=_FIRST_COL + 3)
    if last_data_row >= first_data_row:
        ws.auto_filter.ref = (
            f"{get_column_letter(_FIRST_COL)}{header_row}:"
            f"{get_column_letter(last_col)}{last_data_row}"
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
